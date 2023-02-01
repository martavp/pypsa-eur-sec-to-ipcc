# -*- coding: utf-8 -*-
"""
Script to convert networks from PyPSA-Eur-Sec v0.0.2 to data format used in the
IPCC AR6 database
"""

import pypsa
import openpyxl
import pandas as pd
import yaml
from itertools import product

MWh2EJ=3.6e-9     #convert MWh to EJ
t2Mt=1e-6       # convert tonnes to Mtonnes
MW2GW=0.001       # convert MW/MWh to GW/GWh

# Read the pypsa-eur-sec config file
with open('config.yaml') as yamlfile:
    config = yaml.load(yamlfile,Loader=yaml.FullLoader)

#original IPCC file, official template
template_path = "global_sectoral/IPCC_AR6_WG3_Global_sectoral_Pathways_scenario_template_v3.1_20221027.xlsx"
#use the official template

# Metadata
model = "PyPSA-Eur-Sec"
model_version = config['version']
literature_reference = "Pedersen, T. T., Gøtske, E. K., Dvorak, A., Andresen, G. B., & Victoria, M. (2022). Long-term implications of reduced gas imports on the decarbonization of the European energy system. Joule, 6(7), 1566-1580."
sector_opts = config['scenario']['sector_opts']
h=int(sector_opts[0][ : str.find(sector_opts[0], 'H')]) 


keys, values = zip(*config['scenario'].items())
permutations_dicts = [dict(zip(keys, v)) for v in product(*values)]
scenarios = []
for scenario_i in permutations_dicts:
    scenarios.append("elec_s{simpl}_{clusters}_lv{lv}_{opts}_{sector_opts}_".format(**scenario_i))

output_folder = 'results/'

years = config['scenario']['planning_horizons']

countries = ['AT','BE','BG','CH','CZ','DE','DK','EE','ES','FI','FR','GB','GR','HR',
             'HU','IT','LT','LU','LV','NO','PL','PT','RO','SE','SI','SK','IE', 'NL',
             #'RS','BA'
             ] 

iso2name={'AT':'Austria',
          'BE':'Belgium',
          'BG':'Bulgaria',
          'CH':'Switzerland',
          'CZ':'Czech Republic',
          'DE':'Germany',
          'DK':'Denmark',
          'EE':'Estonia',
          'ES':'Spain',
          'FI':'Finland',
          'FR':'France',
          'GB':'United Kingdom',
          'GR':'Greece',
          'HR':'Croatia',
          'HU':'Hungary',
          'IT':'Italy',
          'LT':'Lithuania',
          'LU':'Luxembourg',
          'LV':'Latvia',
          'NO':'Norway',
          'PL':'Poland',
          'PT':'Portugal',
          'RO':'Romania',
          'SE':'Sweden',
          'SI':'Slovenia',
          'SK':'Slovakia',
          'IE':'Ireland',
          'NL':'The Netherlands',}

dict_industry= {
   'Cement':['Cement'],
   'Chemicals|Ammonia':['Ammonia'],
   'Chemicals|High value chemicals': ['HVC'],
   'Chemicals|Methanol':['Methanol'],
   'Chemicals|Other': ['Other chemicals'],
   'Non-ferrous metals':['Aluminium - primary production','Aluminium - secondary production','Other non-ferrous metals'],
   'Other': ['Other Industrial Sectors'],
   'Pulp and Paper' : ['Pulp production','Paper production'],
   'Steel' :['Electric arc','DRI + Electric arc','Integrated steelworks']}

industry_inputs= {
    'Electricity':['elec'],
    'Gases|Fossil':['methane'],
    'Heat':['heat'],
    'Hydrogen':['hydrogen'],
    'Liquids|Fossil':['naphtha'],
    'Solids':['coal','coke','biomass']}

for scenario in scenarios:
    #one excel file per scenario
    file = openpyxl.load_workbook(template_path)
    ds = file['data'] #data sheet
    ratios = pd.read_csv("resources/industry_sector_ratios.csv",index_col=0,header=0,)

    for year in years:
        n = pypsa.Network(f"postnetworks/{scenario}{year}.nc")
        costs = pd.read_csv(f"costs/costs_{year}.csv", index_col=[0,1])
        prod = pd.read_csv("resources/industrial_production_{}_{}.csv".format(scenario[ : str.find(scenario, '_lv')]
                                                   ,year),index_col=0,header=0,)  
        industry_demand=pd.read_table("resources/industrial_energy_demand_{}_{}.csv".format(scenario[ : str.find(scenario, '_lv')]
                                                   ,year),delimiter=',',index_col=0)
        col=[c for c in ds[1] if c.value==year][0].column
        
        for i,country in enumerate(countries):
            if year == years[0]:
                #one datasheet per country including information from different years
                target = file.copy_worksheet(file['data'])
                target.title ='data' + str(i)
            ds = file['data' + str(i)] 
            var={}
            
            dict_var={'Capacity':' ', 'Capacity Additions':year}      
            for v_type in dict_var.keys():
               """
               Capacity : Solar PV, onshore and offshore wind
               """
               #MW -> GW
               var[v_type+'|Electricity|Solar|PV'] =MW2GW*(n.generators.p_nom_opt.filter(like ='solar').filter(like =country).filter(like=str(dict_var[v_type])).sum()-
                                                          n.generators.p_nom_opt.filter(like ='solar thermal').filter(like =country).filter(like=str(dict_var[v_type])).sum())  # Electricity|Solar|PV : thermal is deducted since it's used to produce heat
               var[v_type+'|Electricity|Solar'] = var[v_type+'|Electricity|Solar|PV']
               var[v_type+'|Electricity|Solar|PV|Rooftop PV'] = MW2GW*n.generators.p_nom_opt.filter(like ='solar rooftop').filter(like =country).filter(like=str(dict_var[v_type])).sum()
               var[v_type+'|Electricity|Solar|PV|Utility-scale PV'] = var[v_type+'|Electricity|Solar|PV'] - var[v_type+'|Electricity|Solar|PV|Rooftop PV']
   
               var[v_type+'|Electricity|Wind|Onshore']=MW2GW*n.generators.p_nom_opt.filter(like ='onwind').filter(like =country).filter(like=str(dict_var[v_type])).sum() 
               var[v_type+'|Electricity|Wind|Offshore']=MW2GW*n.generators.p_nom_opt.filter(like ='offwind').filter(like =country).filter(like=str(dict_var[v_type])).sum() 
               var[v_type+'|Electricity|Wind']=var[v_type+'|Electricity|Wind|Onshore']+var[v_type+'|Electricity|Wind|Offshore']


               """
                Capacity : Nuclear, Coal, Lignite, OCGT, CCGT, Biomass, Oil
               """
               #MW -> GW
               var[v_type+'|Electricity|Nuclear'] =MW2GW*((n.links.efficiency.filter(like ='nuclear').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='nuclear').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
            
               var[v_type+'|Electricity|Coal|w/o CCS'] = ((n.links.efficiency.filter(like ='coal').filter(like =country).filter(like=str(dict_var[v_type]))
                *n.links.p_nom_opt.filter(like ='coal').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               var[v_type+'|Electricity|Coal|w/o CCS'] += MW2GW*((n.links.efficiency.filter(like ='lignite').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='lignite').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               var[v_type+'|Electricity|Coal'] =var[v_type+'|Electricity|Coal|w/o CCS'] 
            
               var[v_type+'|Electricity|Gas'] = MW2GW*((n.links.efficiency.filter(like ='OCGT').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='OCGT').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               var[v_type+'|Electricity|Gas'] += MW2GW*((n.links.efficiency.filter(like ='CCGT').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='CCGT').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               var[v_type+'|Electricity|Gas'] += MW2GW*((n.links.efficiency.filter(like ='gas CHP').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='gas CHP').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               var[v_type+'|Electricity|Gas|w/ CCS'] = MW2GW*((n.links.efficiency.filter(like ='gas CHP CC').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='gas CHP CC').filter(like =country).filter(like=str(dict_var[v_type]))).sum() )                                                 
               var[v_type+'|Electricity|Gas|w/o CCS'] = (var[v_type+'|Electricity|Gas'] -
                                         var[v_type+'|Electricity|Gas|w/ CCS']  )                                              
                                                              
               var[v_type+'|Electricity|Biomass'] = MW2GW*((n.links.efficiency.filter(like ='solid biomass CHP').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='solid biomass CHP').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               var[v_type+'|Electricity|Biomass|w/ CCS']= MW2GW*((n.links.efficiency.filter(like ='solid biomass CHP CC').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='solid biomass CHP CC').filter(like =country).filter(like=str(dict_var[v_type]))).sum())   
               var[v_type+'|Electricity|Biomass|w/o CCS'] = (var[v_type+'|Electricity|Biomass']    
              -var[v_type+'|Electricity|Biomass|w/ CCS'])
            
               var[v_type+'|Electricity|Oil'] = MW2GW*((n.links.efficiency.filter(like ='oil-').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='oil-').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
            
               """
               Capacity : hydro (reservoir, ror)
               """
               #MW -> GW
               var[v_type+'|Electricity|Hydro'] = MW2GW*n.generators.p_nom_opt.filter(like ='ror').filter(like =country).filter(like=str(dict_var[v_type])).sum()
               var[v_type+'|Electricity|Hydro'] += MW2GW*n.storage_units.p_nom_opt.filter(like ='hydro').filter(like =country).filter(like=str(dict_var[v_type])).sum()
               """
               Capacity : storage (PHS, battery, H2 storage, thermal storage)
               """
               #MWh to GWh
               var[v_type+'|Electricity|Storage|Pumped Hydro Storage'] = MW2GW*(n.storage_units.p_nom_opt.filter(like ='PHS').filter(like =country).filter(like=str(dict_var[v_type]))*
                                                                               n.storage_units.max_hours.filter(like ='PHS').filter(like =country).filter(like=str(dict_var[v_type]))).sum()
               var[v_type+'|Electricity|Storage|Battery Capacity|Home Battery'] =MW2GW*((n.links.efficiency.filter(like ='home battery charger').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='home battery charger').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               var[v_type+'|Electricity|Storage|Battery Capacity'] = MW2GW*((n.links.efficiency.filter(like ='battery charger').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='battery charger').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               var[v_type+'|Electricity|Storage|Battery Capacity|Utility-scale Battery'] = (var[v_type+'|Electricity|Storage|Battery Capacity'] 
                 - var[v_type+'|Electricity|Storage|Battery Capacity|Home Battery'])
               #var[v_type+'|Electricity|Storage|Hydrogen Storage Capacity|overground'] = MW2GW *(n.stores.e_nom_opt.filter(like ='H2').filter(like =country).filter(like=str(dict_var[v_type])).sum()/168) #assume one week charge time for H2 storage
               #var[v_type+'|Electricity|Storage|Hydrogen Storage Capacity|underground'] = MW2GW *n.stores.e_nom_opt[country + ' H2 Store underground'] if country + ' H2 Store underground' in n.stores.index else 0
               var[v_type+'|Electricity|Storage|Hydrogen Storage Capacity'] = MW2GW *(n.stores.e_nom_opt.filter(like ='H2').filter(like =country).filter(like=str(dict_var[v_type])).sum())
               
               var[v_type+'|Storage|Thermal Energy Storage|Household storage'] = n.stores.e_nom_opt.filter(like ='rural water tank').filter(like =country).filter(like=str(dict_var[v_type])).sum()/(3*24) 
               var[v_type+'|Storage|Thermal Energy Storage|District heating storage'] = n.stores.e_nom_opt.filter(like ='central water tank').filter(like =country).filter(like=str(dict_var[v_type])).sum()/(180*24)  #3 day for house and 180 day for rural water tank  charge
               var[v_type+'|Storage|Thermal Energy Storage'] = var[v_type+'|Storage|Thermal Energy Storage|Household storage'] + var[v_type+'|Storage|Thermal Energy Storage|District heating storage']
               
               #MW to GW                                                             
               var[v_type+'|Electricity|Storage Capacity'] = MW2GW* (n.storage_units.p_nom_opt.filter(like ='PHS').filter(like =country).filter(like=str(dict_var[v_type])).sum()   #PHS+battery+hydrogen
                                                            + ((n.links.efficiency.filter(like ='battery charger').filter(like =country).filter(like=str(dict_var[v_type]))
                                                                *n.links.p_nom_opt.filter(like ='battery charger').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
                                                            + (n.links.p_nom_opt.filter(like ='Fuel Cell').filter(like =country).filter(like=str(dict_var[v_type])).sum()))                                  
               """
               Capacity : grid, peak , other
               """
               #MWh to GWh
               var[v_type+'|Electricity|Transmissions Grid'] = MW2GW*((n.lines.s_nom_opt[[i for i in n.lines.index if country in n.lines.bus0[i] or country in n.lines.bus1[i]]]).sum()
                     +(n.links.p_nom_opt[[i for i in n.links.index if 'DC' in n.links.carrier[i] and ((country in n.links.bus0[i]) is not (country in n.links.bus1[i]))]]).sum())
               if v_type == 'Capacity':  
                    var[v_type+'|Electricity|Peak Demand'] = MW2GW*(n.loads_t.p[[i for i in n.loads.index if country in i 
                     and not any(sector in i for sector in ['indust','ship','tran','heat'])]]).max().sum()
                    var[v_type+'|Electricity|Other'] = MW2GW*((n.links.efficiency.filter(like ='H2 Fuel Cell').filter(like =country).filter(like=str(dict_var[v_type]))
                    *n.links.p_nom_opt.filter(like ='H2 Fuel Cell').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               """
               Capacity : heat pumps, heat resistors, Sabatier (synthetic gas)
               """
               # ELectric capacity
               # MW to Gw
               if v_type == 'Capacity': 
                 var[v_type+'|Heating|Heat pumps'] = MW2GW*((n.links_t.efficiency.filter(like ='heat pump').filter(like =country).filter(like=str(dict_var[v_type])).mean()
                    *n.links.p_nom_opt.filter(like ='heat pump').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
                 var[v_type+'|Heating|Electric boilers'] = MW2GW*((n.links.efficiency.filter(like ='resistive heater').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='resistive heater').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
               
               """
               Capacity : Hydrogen and fuel production
               """
               if v_type == 'Capacity':
                var[v_type+'|Hydrogen|Electricity'] =  MW2GW*((n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country).filter(like=str(dict_var[v_type]))
                  *n.links.p_nom_opt.filter(like ='H2 Electrolysis').filter(like =country).filter(like=str(dict_var[v_type]))).sum())       
                var[v_type+'|Hydrogen|Gas|w/ CCS'] =  MW2GW*((n.links.efficiency.filter(like ='SMR CC').filter(like =country).filter(like=str(dict_var[v_type]))
                  *n.links.p_nom_opt.filter(like ='SMR CC').filter(like =country).filter(like=str(dict_var[v_type]))).sum())                                                                       
                var[v_type+'|Hydrogen|Gas'] =  MW2GW*((n.links.efficiency.filter(like ='SMR').filter(like =country).filter(like=str(dict_var[v_type]))
                  *n.links.p_nom_opt.filter(like ='SMR').filter(like =country).filter(like=str(dict_var[v_type]))).sum()) 
                var[v_type+'|Hydrogen|Gas|w/o CCS'] = var[v_type+'|Hydrogen|Gas'] -var[v_type+'|Hydrogen|Gas|w/ CCS'] 
                var[v_type+'|Liquids|Gas'] = MW2GW*((n.links.efficiency.filter(like ='Fischer-Tropsch').filter(like =country).filter(like=str(dict_var[v_type]))
                  *n.links.p_nom_opt.filter(like ='Fischer-Tropsch').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
                var[v_type+'|Gas|Synthetic'] = MW2GW*((n.links.efficiency.filter(like ='Sabatier').filter(like =country).filter(like=str(dict_var[v_type]))
                 *n.links.p_nom_opt.filter(like ='Sabatier').filter(like =country).filter(like=str(dict_var[v_type]))).sum())
            """
            Electricity production : Solar PV, onshore and offshore wind
            """
            #MWh -> EJ
            var['Secondary Energy|Electricity|Solar|PV'] = MWh2EJ*h*(n.generators_t.p.filter(like ='solar').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Solar'] = var['Secondary Energy|Electricity|Solar|PV']
            var['Secondary Energy|Electricity|Solar|PV|Rooftop PV'] = MWh2EJ*h*(n.generators_t.p.filter(like ='solar rooftop').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Solar|PV|Utility-scale PV'] = var['Secondary Energy|Electricity|Solar|PV'] - var['Secondary Energy|Electricity|Solar|PV|Rooftop PV']
            var['Secondary Energy|Electricity|Wind|Onshore'] = MWh2EJ*h*(n.generators_t.p.filter(like ='onwind').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Wind|Offshore'] = MWh2EJ*h*(n.generators_t.p.filter(like ='offwind').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Wind'] = var['Secondary Energy|Electricity|Wind|Onshore'] + var['Secondary Energy|Electricity|Wind|Offshore']
            
            """
            Electricity production : Nuclear, Coal, Lignite, OCGT, CCGT, biomass
            """
            #MWh -> EJ
            var['Secondary Energy|Electricity|Nuclear'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='nuclear').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Coal|w/o CCS'] =- MWh2EJ*h*(n.links_t.p1.filter(like ='coal').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Coal|w/o CCS'] += -MWh2EJ*h*(n.links_t.p1.filter(like ='lignite').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Gas|w/o CCS'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='OCGT').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Gas|w/o CCS'] += -MWh2EJ*h*(n.links_t.p1.filter(like ='CCGT').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Gas|w/o CCS'] += -MWh2EJ*h*(n.links_t.p1.filter(like ='gas CHP').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Gas|w/ CCS'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='gas CHP CC').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Gas|w/o CCS'] -= var['Secondary Energy|Electricity|Gas|w/ CCS']
            var['Secondary Energy|Electricity|Biomass|w/o CCS'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='biomass CHP').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Biomass|w/ CCS'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='biomass CHP CC').filter(like =country).sum().sum())           
            var['Secondary Energy|Electricity|Biomass|w/o CCS'] -= var['Secondary Energy|Electricity|Biomass|w/ CCS']
            """
            Electricity production : Hydro (reservoir, ror)
            """
            #MWh -> EJ
            var['Secondary Energy|Electricity|Hydro'] = MWh2EJ*h*(n.storage_units_t.p.filter(like ='hydro').filter(like =country).sum().sum())
            var['Secondary Energy|Electricity|Hydro'] += MWh2EJ*h*(n.generators_t.p.filter(like ='ror').filter(like =country).sum().sum())
            
            """
            Heat production : Solar, Gas, Biomass
            """
            #MWh -> EJ
            var['Secondary Energy|Heat|Biomass'] = -MWh2EJ*h*(n.links_t.p2.filter(like ='solid biomass CHP').filter(like =country).sum().sum())
            var['Secondary Energy|Heat|Gas'] = -MWh2EJ*h*(n.links_t.p2.filter(like ='gas CHP').filter(like =country).sum().sum()+
                                                            n.links_t.p1.filter(like ='gas boiler').filter(like =country).sum().sum())
            var['Secondary Energy|Heat|Oil'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='oil boiler').filter(like =country).sum().sum())
            var['Secondary Energy|Heat|Solar'] = -MWh2EJ*h*(n.generators_t.p.filter(like='solar thermal').filter(like=country).sum().sum())    
            var['Secondary Energy|Heat|Other'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='resistive heater').filter(like =country).sum().sum()+
                                                            n.links_t.p1.filter(like ='heat pump').filter(like =country).sum().sum())     
            """
            Hydrogen and Synthetic oil production
            """
            #MWh to EJ
            var['Secondary Energy|Hydrogen|Electricity'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='H2 Electrolysis').filter(like =country).sum().sum())
            var['Secondary Energy|Hydrogen|Gas'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='SMR').filter(like =country).sum().sum())
            var['Secondary Energy|Hydrogen|Gas|w/ CCS'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='SMR CC').filter(like =country).sum().sum())
            var['Secondary Energy|Hydrogen|Gas|w/o CCS'] = var['Secondary Energy|Hydrogen|Gas'] - var['Secondary Energy|Hydrogen|Gas|w/ CCS']
            var['Secondary Energy|Liquids|Gas'] = -MWh2EJ*h*(n.links_t.p1.filter(like ='Fischer-Tropsch').filter(like =country).sum().sum())
            
            """
            Final Energy 
            """
            #MWh -> EJ
            
            for industry in dict_industry.keys():
                for industry_name in dict_industry[industry]:
                     for v_type in industry_inputs.keys():
                         for i, v_name in enumerate(industry_inputs[v_type]):
                             if i==0 :  var['Final Energy|Industry|'+industry+'|'+v_type]=0   #skips electricity
                             if 'industry_name' in prod.columns:   #skips some variables that are in newer networks
                                 var['Final Energy|Industry|'+industry+'|'+v_type]+= (   
                                 ratios[industry_name][v_name]*prod[industry_name].filter(like=country).sum)   
  
            var['Final Energy|Industry|Electricity']=MWh2EJ*h*(n.loads_t.p.filter(like ='industry electricity').filter(like =country).sum().sum())    
            var['Final Energy|Industry|Gases|Fossil'] = MWh2EJ*industry_demand['methane'].filter(like=country).sum()  
            var['Final Energy|Industry|Heat']=MWh2EJ*h*(n.loads_t.p.filter(like ='low-temperature heat for industry').filter(like =country).sum().sum())
            var['Final Energy|Industry|Hydrogen']=MWh2EJ*h*(n.loads_t.p.filter(like ='H2 for industry').filter(like =country).sum().sum())
            var['Final Energy|Industry|Liquids|Oil']= MWh2EJ*industry_demand['naphtha'].filter(like=country).sum()
            var['Final Energy|Industry|Solids|Biomass'] =MWh2EJ*industry_demand['solid biomass'].filter(like=country).sum()
            var['Final Energy|Industry|Solids|Coal']=MWh2EJ*(industry_demand['coke'].filter(like=country).sum()
                                        + industry_demand['coal'].filter(like=country).sum())

            var['Final Energy|Residential and Commercial|Electricity'] = MWh2EJ*h*(n.loads_t.p
                       [[i for i in n.loads.index if country in i and not 
                      any(s in i for s in ('transport', 'ship', 'industry', 'agri', 'heat'))]]).sum().sum()
            var['Final Energy|Residential and Commercial|Heating'] =  MWh2EJ*h*(n.loads_t.p.filter(like ='urban').filter(like='heat').filter(like =country).sum().sum()
                          +n.loads_t.p.filter(like ='rural').filter(like='heat').filter(like =country).sum().sum())   
            #var['Final Energy|Transportation|Aviation']= MWh2EJ*h*n.loads_t.p.filter(like ='kerosene for aviation').sum().sum()  for all countries
            var['Final Energy|Transportation|Maritime'] = MWh2EJ*h*(n.loads_t.p.filter(like ='shipping oil').filter(like =country).sum().sum()
                        +n.loads_t.p.filter(like ='H2 for shipping').filter(like =country).sum().sum())
            var['Final Energy|Transportation|Passenger|Electricity'] =MWh2EJ*h*(n.loads_t.p.filter(like ='land transport EV').filter(like =country).sum().sum())    
            var['Final Energy|Transportation|Passenger|Hydrogen']= MWh2EJ*h*(n.loads_t.p.filter(like ='land transport fuel cell').filter(like =country).sum().sum())
            var['Final Energy|Transportation|Passenger|Liquids|Oil'] =MWh2EJ*h*(n.loads_t.p.filter(like ='land transport oil').filter(like =country).sum().sum())


            var['Final Energy|Electricity'] =(var['Final Energy|Residential and Commercial|Electricity']+
                                              var['Final Energy|Industry|Electricity']+
                                              var['Final Energy|Transportation|Passenger|Electricity'])
            var['Final Energy|Heat'] =(var['Final Energy|Residential and Commercial|Heating']+
                                       var['Final Energy|Industry|Heat'])
            var['Final Energy|Hydrogen'] =(var['Final Energy|Industry|Hydrogen']+
                                           var['Final Energy|Transportation|Passenger|Hydrogen'])
            var['Final Energy|Industry']  =  (var['Final Energy|Industry|Electricity']+
            var['Final Energy|Industry|Gases|Fossil'] +var['Final Energy|Industry|Heat']+
                    var['Final Energy|Industry|Hydrogen']+ var['Final Energy|Industry|Liquids|Oil']+
                    var['Final Energy|Industry|Solids|Biomass'] + var['Final Energy|Industry|Solids|Coal'] )
            #50/50 services/domestic
            var['Final Energy|Residential and Commercial|Commercial|Heating|Heat pumps'] = - MWh2EJ * (n.links_t.p1.filter(like ='heat pump').filter(like='residentioal').filter(like =country).sum().sum())
            var['Final Energy|Residential and Commercial|Residential|Heating|Heat pumps'] = - MWh2EJ * (n.links_t.p1.filter(like ='heat pump').filter(like='services').filter(like =country).sum().sum())
            var['Final Energy|Residential and Commercial|Commercial|Heating|Electric boilers'] = - MWh2EJ * (n.links_t.p1.filter(like ='resistive heater').filter(like='services').filter(like =country).sum().sum())
            var['Final Energy|Residential and Commercial|Residential|Heating|Electric boilers'] = - MWh2EJ * (n.links_t.p1.filter(like ='resistive heater').filter(like='residentioal').filter(like =country).sum().sum())
            var['Final Energy|Gas|Synthetic'] = - MWh2EJ*n.links_t.p1.filter(like ='Sabatier').filter(like =country).sum().sum()
            
            """
            Capital cost and Lifetime 
            """
            # €_2015/kW  to US$_2010/kW
            EUR2015_USD2010 = 1.11 /1.09
            ipcc2pypsa ={'Capital Cost': 'investment',
                         'Lifetime':'lifetime'} 
            for metric in ['Capital Cost', 'Lifetime']:
                factor = EUR2015_USD2010 if metric =='Capital Cost' else 1
                var[metric + ' |Electricity|Solar|PV|Rooftop PV'] = factor * costs.loc[('solar-rooftop', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Solar|PV|Utility-scale PV'] = factor * costs.loc[('solar-utility', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Solar|PV'] = 0.5* var[metric + ' |Electricity|Solar|PV|Rooftop PV'] + 0.5*var[metric + '|Electricity|Solar|PV|Utility-scale PV']
                var[metric + '|Electricity|Wind|Onshore'] = factor * costs.loc[('onwind', ipcc2pypsa[metric]), 'value']
                var[metric + '|Electricity|Wind|Offshore'] = factor * costs.loc[('offwind', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Nuclear'] = factor * costs.loc[('nuclear', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Coal|w/o CCS'] = factor * costs.loc[('coal', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Gas|w/o CCS'] = factor * costs.loc[('OCGT', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Biomass|w/o CCS'] = factor * costs.loc[('biomass EOP', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Hydro'] = factor * costs.loc[('hydro', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Storage|Pumped Hydro Storage'] = factor * costs.loc[('PHS', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Storage|Battery Capacity|Utility-scale Battery'] = factor * costs.loc[('battery storage', ipcc2pypsa[metric]),'value']
                var[metric + '|Electricity|Storage|Battery Capacity'] = var [metric + '|Electricity|Storage|Battery Capacity|Utility-scale Battery']
                var[metric + '|Heating|Heat pumps'] = factor * costs.loc[('decentral air-sourced heat pump', ipcc2pypsa[metric]),'value']
                var[metric + '|Heating|Electric boilers'] = factor * costs.loc[('decentral resistive heater', ipcc2pypsa[metric]),'value']
                var[metric + '|Gas|Synthetic'] = factor * costs.loc[('methanation', ipcc2pypsa[metric]),'value']
                var[metric + '|Storage|Thermal Energy Storage|Household storage'] = factor * costs.loc[('decentral water tank storage', ipcc2pypsa[metric]),'value']
                var[metric + '|Storage|Thermal Energy Storage|District heating storage'] = factor * costs.loc[('central water tank storage', ipcc2pypsa[metric]),'value']
                var[metric + '|Hydrogen|Electricity'] = factor * costs.loc[('electrolysis', ipcc2pypsa[metric]),'value']
                var[metric + '|Hydrogen|Gas|w/o CCS'] = factor * costs.loc[('SMR', ipcc2pypsa[metric]),'value']
                var[metric + '|Hydrogen|Gas|w/ CCS'] = factor * costs.loc[('SMR CC', ipcc2pypsa[metric]),'value']
                var[metric + '|Liquids|Gas|w/o CCS'] = factor * costs.loc[('Fischer-Tropsch', ipcc2pypsa[metric]),'value']
        
            """
            OM Cost
            """
            #US$_2010/kW·year
            factor = EUR2015_USD2010
            var['OM Cost |Electricity|Solar|PV|Rooftop PV'] = factor * 0.01*costs.loc[('solar-rooftop', 'FOM'),'value']*costs.loc[('solar-rooftop', 'investment'),'value']
            var['OM Cost|Electricity|Solar|PV|Utility-scale PV'] = factor * 0.01*costs.loc[('solar-utility', 'FOM'),'value']*costs.loc[('solar-utility', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Solar|PV'] = 0.5* var['OM Cost |Electricity|Solar|PV|Rooftop PV'] + 0.5*var['OM Cost|Electricity|Solar|PV|Utility-scale PV']
            var['OM Cost|Fixed|Electricity|Wind|Onshore'] = factor * 0.01*costs.loc[('onwind', 'FOM'),'value']*costs.loc[('onwind', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Wind|Offshore'] = factor * 0.01*costs.loc[('offwind', 'FOM'),'value']*costs.loc[('offwind', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Nuclear'] = factor * 0.01*costs.loc[('nuclear', 'FOM'),'value']*costs.loc[('nuclear', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Coal|w/o CCS'] = factor * 0.01*costs.loc[('coal', 'FOM'),'value']*costs.loc[('coal', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Gas|w/o CCS'] = factor * 0.01*costs.loc[('OCGT', 'FOM'),'value']*costs.loc[('OCGT', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Hydro'] = factor * 0.01*costs.loc[('hydro', 'FOM'),'value']*costs.loc[('hydro', 'investment'),'value']
            var['OM Cost|Electricity|Storage|Pumped Hydro Storage'] = factor * 0.01*costs.loc[('PHS', 'FOM'),'value']*costs.loc[('PHS', 'investment'),'value']
            var['OM Cost|Electricity|Storage|Battery Capacity|Utility-scale Battery '] = factor * 0.01*costs.loc[('battery storage', 'FOM'),'value']*costs.loc[('battery storage', 'investment'),'value']
            var['OM Cost|Electricity|Storage|Battery Capacity'] = var [metric + '|Electricity|Storage|Battery Capacity|Utility-scale Battery']
            # The below two lines are outcommented due to error in the EU Climate advisory board side. 
            # The lines should be included again in the future
            var['OM Cost|Electricity|Storage|Hydrogen Storage Capacity|Overground'] = factor * 0.01*costs.loc[('hydrogen storage tank', 'FOM'),'value']*costs.loc[('hydrogen storage tank', 'investment'),'value']
            var['OMCost|Electricity|Storage|Hydrogen Storage Capacity|Underground'] = factor * 0.01*costs.loc[('hydrogen storage underground', 'FOM'),'value']*costs.loc[('hydrogen storage underground', 'investment'),'value']
            var['OM Cost|Heating|Heat pumps'] =factor * 0.01*costs.loc[('decentral air-sourced heat pump', 'FOM'),'value']*costs.loc[('decentral air-sourced heat pump', 'investment'),'value']
            var['OM Cost|Heating|Electric boilers'] = factor * 0.01*costs.loc[('decentral resistive heater', 'FOM'),'value']*costs.loc[('decentral resistive heater', 'investment'),'value']
            var['OM Cost|Gas|Synthetic'] = factor * 0.01*costs.loc[('methanation', 'FOM'),'value']*costs.loc[('methanation', 'investment'),'value']
            var['OM Cost|Storage|Thermal Energy Storage|Household storage'] = factor * 0.01*costs.loc[('decentral water tank storage', 'FOM'),'value']*costs.loc[('decentral water tank storage', 'investment'),'value']
            var['OM Cost|Storage|Thermal Energy Storage|District heating storage'] = factor * 0.01*costs.loc[('central water tank storage', 'FOM'),'value']*costs.loc[('central water tank storage', 'investment'),'value']
            var['OM Cost|Fixed|Hydrogen|Electricity'] = factor * 0.01*costs.loc[('electrolysis', 'FOM'),'value']*costs.loc[('electrolysis', 'investment'),'value']
            var['OM Cost|Fixed|Electricity|Biomass|w/o CCS'] = factor * 0.01*costs.loc[('biomass EOP', 'FOM'),'value']*costs.loc[('biomass EOP', 'investment'),'value']

            """
            Investment
            """
            var['Investment|Energy Supply|Electricity|Biomass|w/ CCS'] = (n.links.p_nom_opt.filter(like='biomass CHP CC').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='biomass CHP CC').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Electricity|Biomass|w/o CCS'] = (n.links.p_nom_opt.filter(like='biomass CHP').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='biomass CHP').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Electricity|Coal|w/ CCS'] = (n.links.p_nom_opt.filter(like='coal CC').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='coal CC').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Electricity|Coal|w/o CCS'] = (n.links.p_nom_opt.filter(like='coal').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='coal').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Electricity|Electricity Storage'] = ((n.links.p_nom_opt.filter(like='battery').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='battery').filter(like=str(year)).filter(like=country)).sum()+
                                         (n.stores.e_nom_opt.filter(like='battery').filter(like=str(year)).filter(like=country)*
                                         n.stores.capital_cost.filter(like='battery').filter(like=str(year)).filter(like=country)).sum())
            var['Investment|Energy Supply|Electricity|Gas'] = (n.links.p_nom_opt.filter(like='CGT').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='CGT').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Electricity|Fossil'] = (var['Investment|Energy Supply|Electricity|Coal|w/ CCS']
                                                               + var['Investment|Energy Supply|Electricity|Coal|w/o CCS']
                                                               + var['Investment|Energy Supply|Electricity|Gas'] )
            var['Investment|Energy Supply|Electricity|Hydro'] =((n.generators.p_nom_opt.filter(like='ror').filter(like=str(year)).filter(like=country)*
                                         n.generators.capital_cost.filter(like='ror').filter(like=str(year)).filter(like=country)).sum()+
                                         (n.links.p_nom_opt.filter(like='hydro').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='hydro').filter(like=str(year)).filter(like=country)).sum())          
            var['Investment|Energy Supply|Electricity|Nuclear'] = (n.links.p_nom_opt.filter(like='nuclear').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='nuclear').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Electricity|Other'] = ((n.links.p_nom_opt.filter(like='H2 Fuel Cell').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='H2 Fuel Cell').filter(like=str(year)).filter(like=country)).sum())
            var['Investment|Energy Supply|Electricity|Solar'] =(n.generators.p_nom_opt.filter(like='solar').filter(like=str(year)).filter(like=country)*
                                         n.generators.capital_cost.filter(like='solar').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Electricity|Transmission and Distribution'] = ((n.links.p_nom_opt.filter(like='distribution grid').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='distribution grid').filter(like=str(year)).filter(like=country)).sum()+
             (n.lines.capital_cost*n.lines.s_nom_opt)[[i for i in n.lines.index if country in n.lines.bus0[i] or country in n.lines.bus1[i]]].sum()+
             (n.links.p_nom_opt*n.links.capital_cost)[[i for i in n.links.index if 'DC' in n.links.carrier[i] and ((country in n.links.bus0[i]) is not (country in n.links.bus1[i]))]].sum())                                                                            
            var['Investment|Energy Supply|Electricity|Wind'] =(n.generators.p_nom_opt.filter(like='wind').filter(like=str(year)).filter(like=country)*
                                         n.generators.capital_cost.filter(like='wind').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Hydrogen|Electricity'] = (n.links.p_nom_opt.filter(like='H2 electrolysis').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='H2 electrolysis').filter(like=str(year)).filter(like=country)).sum()
            var['Investment|Energy Supply|Hydrogen|Fossil'] =  (n.links.p_nom_opt.filter(like='SMR').filter(like=str(year)).filter(like=country)*
                                         n.links.capital_cost.filter(like='SMR').filter(like=str(year)).filter(like=country)).sum()


            """
            Efficiency
            """
            var['Efficiency|Electricity|Biomass|w/ CCS']=n.links.efficiency.filter(like ='solid biomass CHP CC').filter(like =country).mean()
            var['Efficiency|Electricity|Biomass|w/o CCS']=n.links.efficiency.filter(like ='solid biomass CHP').filter(like =country).mean()
            var['Efficiency|Electricity|Coal|w/ CCS']= n.links.efficiency.filter(like ='coal CC').filter(like =country).mean()  
            var['Efficiency|Electricity|Coal|w/o CCS']= n.links.efficiency.filter(like ='coal').filter(like =country).mean()  
            var['Efficiency|Electricity|Gas|w/ CCS']=n.links.efficiency.filter(like ='gas CHP CC').filter(like =country).mean()
            var['Efficiency|Electricity|Gas|w/o CCS']=(n.links.efficiency.filter(like ='gas CHP').filter(like =country).mean()
                                          + n.links.efficiency.filter(like ='OCGT').filter(like =country).mean() 
                                           + n.links.efficiency.filter(like ='CCGT').filter(like =country).mean())/3

            var['Efficiency|Hydrogen|Electricity']=n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country).mean()
            var['Efficiency|Hydrogen|Gas|w/ CCS']=n.links.efficiency.filter(like ='SMR CC').filter(like =country).mean()
            var['Efficiency|Hydrogen|Gas|w/o CCS']=n.links.efficiency.filter(like ='SMR').filter(like =country).mean()
            var['Effciency|Heating|Heat pumps']=(-1)*(n.links_t.p1.filter(like='heat pump').filter(like=country).sum()/
                                    n.links_t.p0.filter(like='heat pump').filter(like=country).sum()).mean()
            var['Efficiency|Heating|Electric boilers']=n.links.efficiency.filter(like='resistive heater').filter(like=country).mean()
            var['Efficiency|Gas|Synthetic']=(n.links.efficiency.filter(like ='Sabatier').filter(like =country).mean()
                 +n.links.efficiency.filter(like ='helmeth').filter(like =country).mean()
                 +n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country).mean())/3
            var['Efficiency|Oil|Synthetic']=n.links.efficiency.filter(like='Fischer-Tropsch').filter(like=country).mean()


            """
            Emissions and Carbon intensity for industry (MtCO2)   #carbon intensity or emissions
            """
             
            
            var['Emissions|CO2|Energy|Supply']=(-1)*t2Mt*(n.links_t.p1[[i for i in n.links.index if 'co2 atm' in n.links.bus1[i] and country in i]].sum().sum()
                                     +n.links_t.p2[[i for i in n.links.index if 'co2 atm' in n.links.bus2[i] and 'nuclear' not in i and country in i]].sum().sum()
                                     +n.links_t.p3[[ i for i in n.links.index if 'co2 atm' in n.links.bus3[i] and country in i]].sum().sum())

            var['Carbon Intensity|Production|Cement']=t2Mt*ratios['Cement']['process emission']*prod['Cement'].filter(like=country).sum()
            var['Carbon Intensity|Production|Chemicals|Ammonia']=t2Mt*ratios['Ammonia']['process emission']*prod['Ammonia'].filter(like=country).sum()
            if 'HVC' in prod.columns: var['Carbon Intensity|Production|Chemicals|High value chemicals']=(t2Mt*ratios['HVC']['process emission']*prod['HVC'].filter(like=country).sum()+
                   t2Mt*ratios['HVC']['process emission from feedstock']*prod['HVC'].filter(like=country).sum())
            if 'Methanol' in prod.columns: var['Carbon Intensity|Production|Chemicals|Methanol']=t2Mt*ratios['Methanol']['process emission']*prod['Methanol'].filter(like=country).sum()
            var['Carbon Intensity|Production|Chemicals|Other']=t2Mt*ratios['Other chemicals']['process emission']*prod['Other chemicals'].filter(like=country).sum()
            var['Carbon Intensity|Production|Non-ferrous metals']=(t2Mt*ratios['Aluminium - primary production']['process emission']*prod['Aluminium - primary production'].filter(like=country).sum()+
                   t2Mt*ratios['Other non-ferrous metals']['process emission']*prod['Other non-ferrous metals'].filter(like=country).sum())/2
            var['Carbon Intensity|Production|Other']=t2Mt*ratios['Other Industrial Sectors']['process emission']*prod['Other Industrial Sectors'].filter(like=country).sum()  
            var['Carbon Intensity|Production|Pulp and Paper']=(t2Mt*ratios['Pulp production']['process emission']*prod['Pulp production'].filter(like=country).sum()+
                  t2Mt*ratios['Paper production']['process emission']*prod['Paper production'].filter(like=country).sum())/2  
            var['Carbon Intensity|Production|Steel']=(t2Mt*ratios['Electric arc']['process emission']*prod['Electric arc'].filter(like=country).sum()
                 +t2Mt*ratios['DRI + Electric arc']['process emission']*prod['DRI + Electric arc'].filter(like=country).sum()
                 +t2Mt*ratios['Integrated steelworks']['process emission']*prod['Integrated steelworks'].filter(like=country).sum())/3
            
            """
            Carbon utilization and sequestration (Mtco2)  ## Mtco2?
            """
            var['Carbon Sequestration|CCS|Biomass']=t2Mt*h*n.links_t.p4.filter(like='biomass CHP CC').filter(like=country).sum().sum()
            var['Carbon Sequestration|CCS|Fossil']=t2Mt*h*(-1)*(n.links_t.p3.filter(like='gas CHP CC').filter(like=country).sum().sum()
                                        +n.links_t.p4.filter(like='gas CHP CC').filter(like=country).sum().sum()) 
            var['Carbon Sequestration|CCS|Industrial Processes']=t2Mt*h*(-1)*n.links_t.p3.filter(like='SMR CC').filter(like=country).sum().sum()
            var['Carbon Sequestration|Direct Air Capture']=t2Mt*h*n.links_t.p0.filter(like='DAC').filter(like=country).sum().sum()
            var['Carbon Sequestration|CCS']=(var['Carbon Sequestration|CCS|Biomass']
                                 +var['Carbon Sequestration|CCS|Fossil']
                                 +var['Carbon Sequestration|CCS|Industrial Processes']
                                 +var['Carbon Sequestration|Direct Air Capture'])
            var['Carbon Utilization|CCS|Industry']= t2Mt*h*(n.links_t.p2.filter(like='Sabatier').filter(like=country).sum().sum()
                                                    +n.links_t.p2.filter(like='helmeth').filter(like=country).sum().sum()
                                                     +n.links_t.p2.filter(like='Fischer-Tropsch').filter(like=country).sum().sum()) 


            for v in var.keys():
                ro=[r for r in ds['D'] if r.value==v][0].row
                ds.cell(row=ro, column=col).value = round(var[v],3) 
                ds.cell(row=ro, column=1).value = model
                ds.cell(row=ro, column=2).value = scenario
                ds.cell(row=ro, column=3).value = iso2name[country] #region
    # add scenario name to 'meta_scenario' sheet
    ds2 = file['meta_scenario']
    ds2.cell(row=4, column=2).value = scenario
    ds2.cell(row=4, column=4).value = model
    ds2.cell(row=4, column=5).value = model_version
    ds2.cell(row=4, column=10).value = literature_reference
    try:
        climate_target = float(scenario[scenario.find('cb')+2:scenario.find('cb')+6])
        ds2.cell(row=4, column=14).value = climate_target
    except: 
        pass

 
        
    file.save(f"{output_folder}IPCC_AR6_{scenario}_new.xlsx")

