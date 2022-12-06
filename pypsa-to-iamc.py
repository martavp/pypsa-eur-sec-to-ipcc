"""
Script to convert networks from PyPSA-Eur-Sec v0.5.0 to data format used in the
IAMC database
"""

import pypsa
import openpyxl
import pandas as pd
import numpy as np
import math
import collections

def select_time_period(series) :
 """
 returns the series sliced in a specified time period (summer, winter,or just all year)
 """    
 if 'Winter'in sheet  :
  a=pd.Series(series[winter_i:winter_e])
  b=pd.Series(series[winter_ii:winter_ee])
  return pd.concat([a,b])
 if 'summer' in sheet or 'Summer' in sheet :         
  return pd.Series(series[summer_i:summer_e]) 
 else :
  return series     

def select_metric(series) :
 """
 returns the peak (max), percentile 25,  percentile 50, or total sum of a series
 """
 if 'peak' in sheet :
  return series.max()
 if 'Percentile' in sheet and '50' in sheet:
  return series.quantile(0.5)
 if 'Percentile' in sheet and '25' in sheet :
  return series.quantile(0.25)
 else :
  return series.sum() 

def safe_div(n, d):
 """
 Only divides if the divisor is not zero, otherwise returns zero.
 """ 
 return n / d if d else 0
#Defining constants (values in pypsa are mainly described in MW)

summer_i='2013-06-01 00:00:00'
summer_e='2013-09-30 23:00:00'
winter_i='2013-11-01 00:00:00'
winter_e='2013-12-31 23:00:00'
winter_ii='2013-01-01 00:00:00'
winter_ee='2013-02-28 23:00:00'

h=3              #hourly resolution
MWh2TJ=3.6e-3     #convert MWh to TJ
TWh2TJ=3.6e+3   #convert TWh to TJ
MW2GW=0.001

#original IAMC file, official template
path = "format/IAMC_format.xlsx"

model = "PyPSA-Eur-Sec 0.6.0"
scenarios ={'test scenario':'-cb25.7ex0',
            #'Climate neutrality scenario':'-cb45.0ex0',
            #'Current trends':''
            }   #defines carbon budget for each scenario
years = [2020]  

sheets={"data_installed_capacity":0,
        "data_fuel_consumption_supply":1,
        "data_Emissions_supply":2,
        "data_Yearly_generation_supply":3,
        "data_Winter_peak_generation":4,
        "data_summer_peak_generation":5,
        "data_Percentile50_generation":6,
        "data_Percentile25_generation":7,
        "data_Investments":8,
        "data_Demand_final_energy":9,
        "data_Demand_emissions":10,
        "data_Percentile_25_consumpt":11,
        "data_Percentile_50_consumpt":12,
        "data_Summer_peak_consumpt":13,
        "data_Winter_peak_consumpt":14,
        "data_Efficiency_demand":15,
        "data_Efficiency_supply":16,
        "data_Other_variables":17}

            
sheet_var={"installed_capacity":'Installed capacity',
           "fuel_consumption_supply":"Fuel consumption",
           "Emissions_supply":"Emissions|Kyoto gases|Fossil|CO2",
           "Yearly_generation_supply":'Generation|Yearly',
           "Winter_peak_generation":'Generation|Winter peak',
           "summer_peak_generation":'Generation|Summer peak',
           "Percentile50_generation":'Generation|Percentile 50',
           "Percentile25_generation":'Generation|Percentile 25',
           'Percentile_25_consumpt':'Hourly power consumption|Percentile 25',
           'Percentile_50_consumpt':'Hourly power consumption|Percentile 50',
           'Summer_peak_consumpt' :'Hourly power consumption|Summer peak',
           'Winter_peak_consumpt':'Hourly power consumption|Winter peak',
           'Investments':'Investments',
           'Demand_final_energy':'Demand_final_energy',
           'Demand_emissions':'Emissions|Kyoto gases|Fossil|CO2',
           'Efficiency_demand':'Efficiency',
           'Efficiency_supply':'Efficiency',
           'Other_variables':'',
           'ENBIOS':''}


for scenario in scenarios:
    #one excel file per scenario
    file = openpyxl.load_workbook(path)
    ds_eu, ds_eu_uk, ds_all=([] for i in range(3))  #separate sheets for regions
    
    for year in years:
        
        n=pypsa.Network('postnetworks/elec_s370_37m_lv1.0__3H-T-H-B-I-solar+p3-dist1{}_{}.nc'.format(scenarios[scenario],year))
        industry_demand=pd.read_table('resources/industrial_energy_demand_elec_s370_37m_{}.csv'.format(year),delimiter=',',index_col=0)
        costs = pd.read_csv("costs/costs_{}.csv".format(year), index_col=[0,1])
        year_sub=2015 if year==years[0] else year    # For the 2015 column of the table, information of 2020 is used.
        countries=list(set(n.buses['country']))
        countries.remove('')
        countries.insert(0,'EU_UK')
        countries.insert(0,'EU')
        EU_UK = [country for country in countries if country not in ('NO','BA','ME','MK','RS','AL','CH','EU')]  #EU25+UK (EU25=EU27-Malta and Cyprus not in model)
        EU = [country for country in countries if country not in ('NO','BA','ME','MK','RS','AL','CH','GB','EU_UK')]  #EU25 (EU27-Malta and Cyprus not in model)
        countries.insert(0,'All')  #All countries being modeled
        All= [country for country in countries if country not in ('EU','EU_UK')]  
        n_eff_eu,n_eff_eu_uk,n_eff_all =(collections.defaultdict(lambda: 0, {}) for i in range(3)) #to keep track of efficiencies for regions
        regions={ 'EU':[EU,ds_eu,n_eff_eu], 'EU_UK':[EU_UK,ds_eu_uk,n_eff_eu_uk], 'All': [All,ds_all,n_eff_all]}

        for country in countries:
          # Prepare ds and var for the specific country 
            ds=[]
            var=[]
            for sheet in sheets.keys() :
              if year==years[0]:
                #one datasheet per country including information from different years
                target = file.copy_worksheet(file[sheet])
                target.title =sheet +' '+str(country)
                if country is 'EU' :  #Separate worksheet for EU so other values could be added as the program runs
                  ds_eu.append(file[sheet +' '+str(country)]) 
                if country is 'EU_UK' :  #Separate worksheet for EU+UK
                  ds_eu_uk.append(file[sheet +' '+str(country)])   
                if country is 'All' :  #Separate worksheet for All counties being modeled
                  ds_all.append(file[sheet +' '+str(country)])   
              ds.append(file[sheet +' '+str(country)])  #set of worksheets for each country that is replaced each loop
              var.append({}) 


            #############################################
            ######  Fill installed capacities sheet #####
            #############################################            
            ###Electric capcaties (MW -> GW)
            sh=sheets['installed_capacity']  #first sheet
            
            #Capacity : Solar PV (rooftop, ground)   
            var[sh]['Installed capacity|Electricity|Solar|Rooftop PV'] = MW2GW*n.generators.p_nom_opt.filter(like ='solar rooftop').filter(like =country).sum()
            var[sh]['Installed capacity|Heat|Solar'] = MW2GW*n.generators.p_nom_opt.filter(like ='solar thermal').filter(like =country).sum()
            var[sh]['Installed capacity|Electricity|Solar|Open field'] = (MW2GW*n.generators.p_nom_opt.filter(like ='solar').filter(like =country).sum()
            -var[sh]['Installed capacity|Heat|Solar']-var[sh]['Installed capacity|Electricity|Solar|Rooftop PV'])
            var[sh]['Installed capacity|Electricity|Solar'] =var[sh]['Installed capacity|Electricity|Solar|Open field']+var[sh]['Installed capacity|Electricity|Solar|Rooftop PV']
            
            #Capacity :  onshore and offshore wind   
            var[sh]['Installed capacity|Electricity|Wind|Onshore']=MW2GW*n.generators.p_nom_opt.filter(like ='onwind').filter(like =country).sum() 
            var[sh]['Installed capacity|Electricity|Wind|Offshore']=MW2GW*n.generators.p_nom_opt.filter(like ='offwind').filter(like =country).sum() 
            var[sh]['Installed capacity|Electricity|Wind']=var[sh]['Installed capacity|Electricity|Wind|Onshore']+var[sh]['Installed capacity|Electricity|Wind|Offshore']
            
            #Capacity : Nuclear
            var[sh]['Installed capacity|Electricity|Nuclear'] =MW2GW*((n.links.efficiency.filter(like ='nuclear').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='nuclear').filter(like =country)).sum())
          

            #Capacity :  Coal (Lignite)                                                                   
            var[sh]['Installed capacity|Electricity|Coal|Brown Coal|Lignite'] = MW2GW*((n.links.efficiency.filter(like ='lignite').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='lignite').filter(like =country)).sum())
            var[sh]['Installed capacity|Electricity|Coal|Brown Coal'] =var[sh]['Installed capacity|Electricity|Coal|Brown Coal|Lignite'] 
            
            var[sh]['Installed capacity|Electricity|Coal'] = MW2GW*((n.links.efficiency.filter(like ='coal').filter(like =country)
                *n.links.p_nom_opt.filter(like ='coal').filter(like =country)).sum())+ var[sh]['Installed capacity|Electricity|Coal|Brown Coal']                                                    

            #Capacity : Natural gas(OCGT, CCGT, CHP, CHP CC)                                                                        
            var[sh]['Installed capacity|Electricity|Gases|Fossil|Natural gas'] = MW2GW*((n.links.efficiency.filter(like ='gas CHP').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='gas CHP').filter(like =country)).sum())
            var[sh]['Installed capacity|Electricity|Gases|Fossil|Natural gas'] += MW2GW*((n.links.efficiency.filter(like ='OCGT').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='OCGT').filter(like =country)).sum())
            var[sh]['Installed capacity|Electricity|Gases|Fossil|Natural gas'] += MW2GW*((n.links.efficiency.filter(like ='CCGT').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='CCGT').filter(like =country)).sum())
            
            var[sh]['Installed capacity|Electricity|Gases|Fossil|Natural gas|CCS'] = MW2GW*((n.links.efficiency.filter(like ='gas CHP CC').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='gas CHP CC').filter(like =country)).sum()) 
                
            #Capacity :Biomass (CCS)                                                                 
            var[sh]['Installed capacity|Electricity|Solid bio and waste|Primary solid biomass'] = MW2GW*((n.links.efficiency.filter(like ='solid biomass CHP').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='solid biomass CHP').filter(like =country)).sum()) 
            var[sh]['Installed capacity|Electricity|Solid bio and waste'] = var[sh]['Installed capacity|Electricity|Solid bio and waste|Primary solid biomass']
            var[sh]['Installed capacity|Electricty|Biomass'] = var[sh]['Installed capacity|Electricity|Solid bio and waste']
            
            var[sh]['Installed capacity|Electricity|Solid bio and waste|Primary solid biomass|CCS'] = MW2GW*((n.links.efficiency.filter(like ='solid biomass CHP CC').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='solid biomass CHP CC').filter(like =country)).sum())            
            var[sh]['Installed capacity|Electricity|Solid bio and waste|CCS'] = var[sh]['Installed capacity|Electricity|Solid bio and waste|Primary solid biomass|CCS']
            var[sh]['Installed capacity|Electricty|Biomass|CCS'] = var[sh]['Installed capacity|Electricity|Solid bio and waste|CCS']            
           
            #Capacity : Hydrogen            
            var[sh]['Installed capacity|Electricity|Gases|Hydrogen'] = MW2GW*((n.links.efficiency.filter(like ='H2 Fuel Cell').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='H2 Fuel Cell').filter(like =country)).sum())

            #Capacity : Oil 
            var[sh]['Installed capacity|Electricity|Liquids|Fossil'] = MW2GW*((n.links.efficiency.filter(like ='oil-').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='oil-').filter(like =country)).sum())

            #Capacity : Power to gas  
            #  According to the SENTINEL team, this variable should include: 
            #  "the power capacity used to produce H2, synthetic methane, synthetic oil or the three of them"            
            var[sh]['Installed capacity|P2G|Electricity'] = MW2GW*((n.links.efficiency.filter(like ='Sabatier').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='Sabatier').filter(like =country)).sum()+ 
                 (n.links.efficiency.filter(like ='helmeth').filter(like =country)*n.links.p_nom_opt.filter(like ='helmeth').filter(like =country)).sum()+ 
                 (n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country)*n.links.p_nom_opt.filter(like ='H2 Electrolysis').filter(like =country)).sum()+
                 (n.links.efficiency.filter(like ='Fischer-Tropsch').filter(like =country)*n.links.p_nom_opt.filter(like ='Fischer-Tropsch').filter(like =country)).sum()) 
            
            #Capacity : hydro (reservoir, ror)
            var[sh]['Installed capacity|Electricity|Hydro|river'] = MW2GW*n.generators.p_nom_opt.filter(like ='ror').filter(like =country).sum()
            var[sh]['Installed capacity|Electricity|Hydro|dam'] = MW2GW*n.storage_units.p_nom_opt.filter(like ='hydro').filter(like =country).sum()
            var[sh]['Installed capacity|Electricity|Hydro'] = var[sh]['Installed capacity|Electricity|Hydro|river']+var[sh]['Installed capacity|Electricity|Hydro|dam']
           
            
            #Capacity : storage (PHS, battery, H2 storage)            
            var[sh]['Installed capacity|Flexibility|Electricity Storage|Medium duration'] = MW2GW*((n.storage_units.p_nom_opt.filter(like ='PHS').filter(like =country).sum()
             +n.stores.e_nom_opt.filter(like ='H2').filter(like =country)).sum()/168)   #assume one week charge time for H2 storage
            
            var[sh]['Installed capacity|Flexibility|Electricity Storage|Short duration'] = MW2GW*(n.links.efficiency.filter(like ='battery charger')
                 *n.links.p_nom_opt.filter(like ='battery charger')).sum()
            # Battery includes utility-scale batteries, home batteries and EV batteries.   
            
            var[sh]['Installed capacity|Flexibility|Electricity Storage'] =  (var[sh]['Installed capacity|Flexibility|Electricity Storage|Medium duration']+
            var[sh]['Installed capacity|Flexibility|Electricity Storage|Short duration'])
            
            #Capacity : Interconnect             
            var[sh]['Installed capacity|Flexibility|Interconnect Importing Capacity'] = MW2GW*((n.lines.s_nom_opt[[i for i in n.lines.index if country in n.lines.bus0[i] or country in n.lines.bus1[i]]]).sum()
            +(n.links.p_nom_opt[[i for i in n.links.index if 'DC' in n.links.carrier[i] and ((country in n.links.bus0[i]) is not (country in n.links.bus1[i]))]]).sum())
                    
                
            ### Heat capcaties (MW -> GW)
            
            #Capacity :Biomass (CCS)
            var[sh]['Installed capacity|Heat|Biomass'] = MW2GW*((n.links.efficiency2.filter(like ='solid biomass CHP').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='solid biomass CHP').filter(like =country)).sum())
            var[sh]['Installed capacity|Heat|Solid bio and waste'] = var[sh]['Installed capacity|Heat|Biomass']  
            var[sh]['Installed capacity|Heat|Solid bio and waste|Primary solid biomass'] = var[sh]['Installed capacity|Heat|Solid bio and waste']  
            
            var[sh]['Installed capacity|Heat|Biomass|CCS'] = MW2GW*((n.links.efficiency2.filter(like ='solid biomass CHP CC').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='solid biomass CHP CC').filter(like =country)).sum())
            var[sh]['Installed capacity|Heat|Solid bio and waste|CCS'] = var[sh]['Installed capacity|Heat|Biomass|CCS']  
            var[sh]['Installed capacity|Heat|Solid bio and waste|Primary solid biomass|CCS'] = var[sh]['Installed capacity|Heat|Solid bio and waste|CCS']  
          
            #Capacity :Electricity (Resistive heater, heat pump)
            var[sh]['Installed capacity|Heat|Electricity|Direct'] = MW2GW*((n.links.efficiency.filter(like ='resistive heater').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='resistive heater').filter(like =country)).sum())
            var[sh]['Installed capacity|Heat|Electricity|Heat pump'] = MW2GW*((n.links_t.efficiency.filter(like ='heat pump').filter(like =country).mean()
                 *n.links.p_nom_opt.filter(like ='heat pump').filter(like =country)).sum())
            var[sh]['Installed capacity|Heat|Electricity']=var[sh]['Installed capacity|Heat|Electricity|Direct']+var[sh]['Installed capacity|Heat|Electricity|Heat pump']

            #Capacity :Natural gas
            var[sh]['Installed capacity|Heat|Gases|Fossil|Natural Gas'] = MW2GW*((n.links.efficiency2.filter(like ='gas CHP').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='gas CHP').filter(like =country)).sum())            
            var[sh]['Installed capacity|Heat|Gases|Fossil|Natural Gas'] += MW2GW*((n.links.efficiency.filter(like ='gas boiler').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='gas boiler').filter(like =country)).sum())
            var[sh]['Installed capacity|Heat|Gases|Fossil|Natural Gas|CCS'] = MW2GW*((n.links.efficiency2.filter(like ='gas CHP CC').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='gas CHP CC').filter(like =country)).sum())              
           
            #Capacity : Oil 
            var[sh]['Installed capacity|Heat|Liquids|Fossil'] = MW2GW*((n.links.efficiency.filter(like ='oil boiler').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='oil boiler').filter(like =country)).sum()) 
                                    
            # Hydrogen capacities 
            var[sh]['Installed capacity|Hydrogen|Electricity'] = MW2GW*((n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='H2 Electrolysis').filter(like =country)).sum())       
            var[sh]['Installed capacity|Hydrogen|Gasses|Fossil|Natural gas|CCS'] = MW2GW*((n.links.efficiency.filter(like ='SMR CC').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='SMR CC').filter(like =country)).sum())                                                                       
            var[sh]['Installed capacity|Hydrogen|Gasses|Fossil|Natural gas'] = MW2GW*((n.links.efficiency.filter(like ='SMR').filter(like =country)
                 *n.links.p_nom_opt.filter(like ='SMR').filter(like =country)).sum())               
                        
            ### Fuel consumption (MWh -> TJ)
            sh=sheets['fuel_consumption_supply']  
            
            ## For Electricity      
            #Fuel : Nuclear
            
            var[sh]['Fuel consumption|Electricity|Nuclear'] =h*MWh2TJ*(n.links_t.p0.filter(like ='nuclear').filter(like =country)).sum().sum()
          
            #Fuel :  Coal (Lignite)        
                                                           
            var[sh]['Fuel consumption|Electricity|Coal|Brown Coal|Lignite'] = h*MWh2TJ*(n.links_t.p0.filter(like ='lignite').filter(like =country)).sum().sum()
            var[sh]['Fuel consumption|Electricity|Coal|Brown Coal'] =var[sh]['Fuel consumption|Electricity|Coal|Brown Coal|Lignite']             
            var[sh]['Fuel consumption|Electricity|Coal'] = h*MWh2TJ*(n.links_t.p0.filter(like ='coal').filter(like =country)).sum().sum() + var[sh]['Fuel consumption|Electricity|Coal|Brown Coal']                                                    

            #Fuel : Natural gas(OCGT, CCGT, CHP, CHP CC)       
               # **CHP fuels calculated bsed on share of electricity and heat, 'safe_div' is used in case thre is division by zero:                                               
            var[sh]['Fuel consumption|Electricity|Gases|Fossil|Natural gas'] = (
                 h*MWh2TJ* (n.links_t.p0.filter(like ='OCGT').filter(like =country)).sum().sum()+
                 h*MWh2TJ*(n.links_t.p0.filter(like ='CCGT').filter(like =country)).sum().sum()+
                 h*MWh2TJ*safe_div(n.links_t.p0.filter(like ='gas CHP').filter(like =country).sum().sum()
                   *n.links_t.p1.filter(like ='gas CHP').filter(like =country).sum().sum()             
                  ,n.links_t.p1.filter(like ='gas CHP').filter(like =country).sum().sum()+
                    n.links_t.p2.filter(like ='gas CHP').filter(like =country).sum().sum()))
            var[sh]['Fuel consumption|Electricity|Gases|Fossil|Natural gas|CCS'] = (
                   h*MWh2TJ*safe_div(n.links_t.p0.filter(like ='gas CHP CC').filter(like =country).sum().sum()
                   *n.links_t.p1.filter(like ='gas CHP CC').filter(like =country).sum().sum()            
                  ,n.links_t.p1.filter(like ='gas CHP CC').filter(like =country).sum().sum()+
                    n.links_t.p2.filter(like ='gas CHP CC').filter(like =country).sum().sum()))
            
            #Fuel :Biomass (CCS)
                                                                 
            var[sh]['Fuel consumption|Electricity|Solid bio and waste|Primary solid biomass'] = (
                h*MWh2TJ*safe_div(n.links_t.p0.filter(like ='solid biomass CHP').filter(like =country).sum().sum()
                  *n.links_t.p1.filter(like ='solid biomass CHP').filter(like =country).sum().sum()            
                  ,n.links_t.p1.filter(like ='solid biomass CHP').filter(like =country).sum().sum()+
                  n.links_t.p2.filter(like ='solid biomass CHP').filter(like =country).sum().sum()))
            var[sh]['Fuel consumption|Electricity|Solid bio and waste'] = var[sh]['Fuel consumption|Electricity|Solid bio and waste|Primary solid biomass']
            var[sh]['Fuel consumption|Electricty|Biomass'] = var[sh]['Fuel consumption|Electricity|Solid bio and waste']
            
            var[sh]['Fuel consumption|Electricity|Solid bio and waste|Primary solid biomass|CCS'] = (
                h*MWh2TJ*safe_div(n.links_t.p0.filter(like ='solid biomass CHP CC').filter(like =country).sum().sum()
                *n.links_t.p1.filter(like ='solid biomass CHP CC').filter(like =country).sum().sum()            
                ,n.links_t.p1.filter(like ='solid biomass CHP CC').filter(like =country).sum().sum()+
                  n.links_t.p2.filter(like ='solid biomass CHP CC').filter(like =country).sum().sum()))            
            var[sh]['Fuel consumption|Electricity|Solid bio and waste|CCS'] = var[sh]['Fuel consumption|Electricity|Solid bio and waste|Primary solid biomass|CCS']
            var[sh]['Fuel consumption|Electricty|Biomass|CCS'] = var[sh]['Fuel consumption|Electricity|Solid bio and waste|CCS']                      
           
            #Fuel : Hydrogen            
            var[sh]['Fuel consumption|Electricity|Gases|Hydrogen'] = h*MWh2TJ*(n.links_t.p0.filter(like ='H2 Fuel Cell').filter(like =country)).sum().sum()

            #Fuel : Oil 
            var[sh]['Fuel consumption|Electricity|Liquids|Fossil'] = h*MWh2TJ*(n.links_t.p0.filter(like ='oil-').filter(like =country)).sum().sum()
            
            #Fuel : Hydrogen 
            var[sh]['Fuel consumption|Hydrogen|Gasses|Fossil|Natural gas'] = h*MWh2TJ*(n.links_t.p0.filter(like ='SMR').filter(like =country)).sum().sum()
            var[sh]['Fuel consumption|Hydrogen|Gasses|Fossil|Natural gas|CCS'] = h*MWh2TJ*(n.links_t.p0.filter(like ='SMR CC').filter(like =country)).sum().sum()
            var[sh]['Fuel consumption|Hydrogen|Electricity'] = h*MWh2TJ*(n.links_t.p0.filter(like ='H2 Electrolysis').filter(like =country)).sum().sum()
                      
            ## For Heat
            #Fuel :Biomass (CCS)
            var[sh]['Fuel consumption|Heat|Biomass'] = (h*MWh2TJ*(n.links_t.p0.filter(like ='solid biomass CHP').filter(like =country)).sum().sum()
               -var[sh]['Fuel consumption|Electricty|Biomass'])
            var[sh]['Fuel consumption|Heat|Solid bio and waste'] = var[sh]['Fuel consumption|Heat|Biomass']  
            var[sh]['Fuel consumption|Heat|Solid bio and waste|Primary solid biomass'] = var[sh]['Fuel consumption|Heat|Solid bio and waste']  
            
            var[sh]['Fuel consumption|Heat|Biomass|CCS'] = (h*MWh2TJ*(n.links_t.p0.filter(like ='solid biomass CHP CC').filter(like =country)).sum().sum()
               -var[sh]['Fuel consumption|Electricty|Biomass|CCS'])
            var[sh]['Fuel consumption|Heat|Solid bio and waste|CCS'] = var[sh]['Fuel consumption|Heat|Biomass|CCS']  
            var[sh]['Fuel consumption|Heat|Solid bio and waste|Primary solid biomass|CCS'] = var[sh]['Fuel consumption|Heat|Solid bio and waste|CCS']  

            #Fuel :Natural gas
            var[sh]['Fuel consumption|Heat|Gases|Fossil|Natural Gas'] = h*MWh2TJ* (n.links_t.p0.filter(like ='gas boiler').filter(like =country)).sum().sum()
            var[sh]['Fuel consumption|Heat|Gases|Fossil|Natural Gas'] += (
                   h*MWh2TJ*safe_div(n.links_t.p0.filter(like ='gas CHP').filter(like =country).sum().sum()
                   *n.links_t.p2.filter(like ='gas CHP').filter(like =country).sum().sum()            
                  ,n.links_t.p1.filter(like ='gas CHP').filter(like =country).sum().sum()+
                    n.links_t.p2.filter(like ='gas CHP').filter(like =country).sum().sum()))
            var[sh]['Fuel consumption|Heat|Gases|Fossil|Natural Gas|CCS'] = (h*MWh2TJ*(n.links_t.p0.filter(like ='gas CHP CC').filter(like =country)).sum().sum() 
              -var[sh]['Fuel consumption|Electricity|Gases|Fossil|Natural gas|CCS'])
                        
            #Fuel : Oil
            var[sh]['Fuel consumption|Heat|Liquids|Fossil'] = h*MWh2TJ*(n.links_t.p0.filter(like ='oil boiler').filter(like =country)).sum().sum()       
                    
            #############################################
            ######  Fill Emissions sheet #####
            #############################################     
            ### Emissions (tCO2 -> MtCO2)
            sh=sheets['Emissions_supply'] 
                  
            #Emissions :  Coal (Lignite)                                                                   
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Coal|Brown Coal|Lignite'] = (1e-6)*(-1)*h*(n.links_t.p2.filter(like ='lignite').filter(like =country)).sum().sum()
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Coal|Brown Coal'] =var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Coal|Brown Coal|Lignite']             
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Coal'] = (1e-6)*(-1)*h*(n.links_t.p2.filter(like ='coal').filter(like =country)).sum().sum() + var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Coal|Brown Coal']                                                    

            #Emissions : Natural gas(OCGT, CCGT, CHP, CHP CC, SMR, SMR CC)       
                                                                 
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Gases|Fossil|Natural gas'] = (
                (1e-6)*(-1)*h*safe_div(n.links_t.p3.filter(like ='gas CHP').filter(like =country).sum().sum()
                 *n.links_t.p1.filter(like ='gas CHP').filter(like =country).sum().sum()             
                 ,n.links_t.p1.filter(like ='gas CHP').filter(like =country).sum().sum()+
                   n.links_t.p2.filter(like ='gas CHP').filter(like =country).sum().sum())) 
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Gases|Fossil|Natural gas|CCS'] = (
                (1e-6)*(-1)*h*safe_div(n.links_t.p3.filter(like ='gas CHP CC').filter(like =country).sum().sum()
                *n.links_t.p1.filter(like ='gas CHP CC').filter(like =country).sum().sum()            
                ,n.links_t.p1.filter(like ='gas CHP CC').filter(like =country).sum().sum()+
                  n.links_t.p2.filter(like ='gas CHP CC').filter(like =country).sum().sum()))
                  
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Gases|Fossil|Natural gas'] += (1e-6)*(-1)*h* (n.links_t.p2.filter(like ='OCGT').filter(like =country)).sum().sum()
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Gases|Fossil|Natural gas'] += (1e-6)*(-1)*h*(n.links_t.p2.filter(like ='CCGT').filter(like =country)).sum().sum()
            
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Hydrogen|Gasses|Fossil|Natural gas'] = (1e-6)*(-1)*h*(n.links_t.p2.filter(like ='SMR').filter(like =country)).sum().sum()
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Hydrogen|Gasses|Fossil|Natural gas|CCS'] = (1e-6)*h*(n.links_t.p2.filter(like ='SMR CC').filter(like =country)).sum().sum()
            
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Gases|Fossil|Natural Gas'] = (
                (1e-6)*(-1)*h*safe_div(n.links_t.p3.filter(like ='gas CHP').filter(like =country).sum().sum()
                *n.links_t.p2.filter(like ='gas CHP').filter(like =country).sum().sum()            
                ,n.links_t.p1.filter(like ='gas CHP').filter(like =country).sum().sum()+
                  n.links_t.p2.filter(like ='gas CHP').filter(like =country).sum().sum()))
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Gases|Fossil|Natural Gas|CCS'] = ((1e-6)*(-1)*h*(n.links_t.p3.filter(like ='gas CHP CC').filter(like =country)).sum().sum()
               -var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Gases|Fossil|Natural gas|CCS'])                                                                              
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Gases|Fossil|Natural Gas'] += (1e-6)*(-1)*h* (n.links_t.p2.filter(like ='gas boiler').filter(like =country)).sum().sum()
            
            #Emissions :Biomass 
            
            #CHP emissions are calculated based on share of electricity and heat:
            
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste|Primary solid biomass'] = (
                (1e-6)*(-1)*h*safe_div(n.links_t.p3.filter(like ='solid biomass CHP').filter(like =country).sum().sum()
               *n.links_t.p1.filter(like ='solid biomass CHP').filter(like =country).sum().sum()            
               ,n.links_t.p1.filter(like ='solid biomass CHP').filter(like =country).sum().sum()+
                 n.links_t.p2.filter(like ='solid biomass CHP').filter(like =country).sum().sum()))
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste'] = var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste|Primary solid biomass']
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricty|Biomass'] = var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste']                          

            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Biomass'] = ((1e-6)*(-1)*h*(n.links_t.p3.filter(like ='solid biomass CHP').filter(like =country)).sum().sum()
               -var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste|Primary solid biomass'])
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Solid bio and waste'] = var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Biomass']  
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Solid bio and waste|Primary solid biomass'] = var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Solid bio and waste']  
            
            #Emissions :Biomass (CCS)
            
            #CHP emissions are calculated based on share of electricity and heat:
            
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste|Primary solid biomass|CCS'] = (
                (1e-6)*(-1)*h*safe_div(n.links_t.p3.filter(like ='solid biomass CHP CC').filter(like =country).sum().sum()
               *n.links_t.p1.filter(like ='solid biomass CHP CC').filter(like =country).sum().sum()            
               ,n.links_t.p1.filter(like ='solid biomass CHP CC').filter(like =country).sum().sum()+
                 n.links_t.p2.filter(like ='solid biomass CHP CC').filter(like =country).sum().sum()))
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste|CCS'] = var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste|Primary solid biomass|CCS']
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricty|Biomass|CCS'] = var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste|CCS']                        

            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Biomass|CCS'] = ((1e-6)*(-1)*h*(n.links_t.p3.filter(like ='solid biomass CHP CC').filter(like =country)).sum().sum()
               -var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Solid bio and waste|Primary solid biomass|CCS'])
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Solid bio and waste|CCS'] = var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Biomass|CCS']  
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Solid bio and waste|Primary solid biomass|CCS'] = var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Solid bio and waste|CCS']  

            #Emissions : Oil 

            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Electricity|Liquids|Fossil'] = (1e-6)*(-1)*h*(n.links_t.p2.filter(like ='oil-').filter(like =country)).sum().sum() 
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Heat|Liquids|Fossil'] = (1e-6)*(-1)*h*(n.links_t.p2.filter(like ='oil boiler').filter(like =country)).sum().sum() 
            
            #############################################
            ######  Fill energy generation sheets #####
            #############################################     
            ###Energy generation: total,peaks,and percentiles (MWh -> GWh)
            for sheet in sheets.keys(): 
             if sheet in ("Yearly_generation_supply","Winter_peak_generation","summer_peak_generation","Percentile50_generation","Percentile25_generation"):
                sh=sheets[sheet]
                
                #Energy generation : Solar PV (rooftop, ground)               
                var[sh][sheet_var[sheet]+'|Electricity|Solar|Rooftop PV'] = MW2GW*h*select_metric(select_time_period(n.generators_t.p.filter(like ='solar rooftop').filter(like =country).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Heat|Solar'] = MW2GW*h*select_metric(select_time_period(n.generators_t.p.filter(like ='solar thermal').filter(like =country).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Electricity|Solar|Open field'] = (MW2GW*h*select_metric(select_time_period(n.generators_t.p.filter(like ='solar').filter(like =country).sum(axis=1)))
                 -var[sh][sheet_var[sheet]+'|Heat|Solar']-var[sh][sheet_var[sheet]+'|Electricity|Solar|Rooftop PV'])
                var[sh][sheet_var[sheet]+'|Electricity|Solar'] =var[sh][sheet_var[sheet]+'|Electricity|Solar|Open field']+var[sh][sheet_var[sheet]+'|Electricity|Solar|Rooftop PV']
             
                #Energy generation :  onshore and offshore wind                
                var[sh][sheet_var[sheet]+'|Electricity|Wind|Onshore']=MW2GW*h*select_metric(select_time_period(n.generators_t.p.filter(like ='onwind').filter(like =country).sum(axis=1))) 
                var[sh][sheet_var[sheet]+'|Electricity|Wind|Offshore']=MW2GW*h*select_metric(select_time_period(n.generators_t.p.filter(like ='offwind').filter(like =country).sum(axis=1))) 
                var[sh][sheet_var[sheet]+'|Electricity|Wind']=var[sh][sheet_var[sheet]+'|Electricity|Wind|Onshore']+var[sh][sheet_var[sheet]+'|Electricity|Wind|Offshore']
             
                #Energy generation : Nuclear (values multiplied by (-1) since pypsa shows generation in links as negative values)            
                var[sh][sheet_var[sheet]+'|Electricity|Nuclear'] =MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='nuclear').filter(like =country)).sum(axis=1)))
          

                #Energy generation :  Coal (Lignite)                                                                   
                var[sh][sheet_var[sheet]+'|Electricity|Coal|Brown Coal|Lignite'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='lignite').filter(like =country)).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Electricity|Coal|Brown Coal'] =var[sh][sheet_var[sheet]+'|Electricity|Coal|Brown Coal|Lignite']          
                var[sh][sheet_var[sheet]+'|Electricity|Coal'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='coal').filter(like =country)).sum(axis=1))) + var[sh][sheet_var[sheet]+'|Electricity|Coal|Brown Coal']                                                    
 
                #Energy generation : Natural gas(OCGT, CCGT, CHP, CHP CC)                                                                        
                var[sh][sheet_var[sheet]+'|Electricity|Gases|Fossil|Natural gas'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='gas CHP').filter(like =country)).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Electricity|Gases|Fossil|Natural gas'] += MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='OCGT').filter(like =country)).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Electricity|Gases|Fossil|Natural gas'] += MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='CCGT').filter(like =country)).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Electricity|Gases|Fossil|Natural gas|CCS'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='gas CHP CC').filter(like =country)).sum(axis=1)))
            
                #Energy generation :Biomass (CCS)                                                                 
                var[sh][sheet_var[sheet]+'|Electricity|Solid bio and waste|Primary solid biomass'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='solid biomass CHP').filter(like =country)).sum(axis=1))) 
                var[sh][sheet_var[sheet]+'|Electricity|Solid bio and waste'] = var[sh][sheet_var[sheet]+'|Electricity|Solid bio and waste|Primary solid biomass']
                var[sh][sheet_var[sheet]+'|Electricty|Biomass'] = var[sh][sheet_var[sheet]+'|Electricity|Solid bio and waste']
            
                var[sh][sheet_var[sheet]+'|Electricity|Solid bio and waste|Primary solid biomass|CCS'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='solid biomass CHP CC').filter(like =country)).sum(axis=1)))            
                var[sh][sheet_var[sheet]+'|Electricity|Solid bio and waste|CCS'] = var[sh][sheet_var[sheet]+'|Electricity|Solid bio and waste|Primary solid biomass|CCS']
                var[sh][sheet_var[sheet]+'|Electricty|Biomass|CCS'] = var[sh][sheet_var[sheet]+'|Electricity|Solid bio and waste|CCS']            
           
                #Energy generation : Hydrogen            
                var[sh][sheet_var[sheet]+'|Electricity|Gases|Hydrogen'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='H2 Fuel Cell').filter(like =country)).sum(axis=1)))

                #Energy generation : Oil 
                var[sh][sheet_var[sheet]+'|Electricity|Liquids|Fossil'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='oil-').filter(like =country)).sum(axis=1)))

                #Energy generation : Power to gas 
                #  According to the SENTINEL team, this variable should include: 
                #  "the power capacity used to produce H2, synthetic methane, synthetic oil or the three of them"
            
                var[sh][sheet_var[sheet]+'|P2G|Electricity'] = MW2GW*(-1)*h*select_metric(select_time_period(n.links_t.p1.filter(like ='Sabatier').filter(like =country).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|P2G|Electricity']+= MW2GW*(-1)*h*select_metric(select_time_period(n.links_t.p1.filter(like ='helmeth').filter(like =country).sum(axis=1))) 
                var[sh][sheet_var[sheet]+'|P2G|Electricity']+= MW2GW*(-1)*h*select_metric(select_time_period(n.links_t.p1.filter(like ='H2 Electrolysis').filter(like =country).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|P2G|Electricity']+= MW2GW*(-1)*h*select_metric(select_time_period(n.links_t.p1.filter(like ='Fischer-Tropsch').filter(like =country).sum(axis=1)))               
              
            
                #Energy generation : hydro (reservoir, ror)
                var[sh][sheet_var[sheet]+'|Electricity|Hydro|river'] = MW2GW*h*select_metric(select_time_period(n.generators_t.p.filter(like ='ror').filter(like =country).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Electricity|Hydro|dam'] = MW2GW*h*select_metric(select_time_period(n.storage_units_t.p.filter(like ='hydro').filter(like =country).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Electricity|Hydro'] = var[sh][sheet_var[sheet]+'|Electricity|Hydro|river']+var[sh][sheet_var[sheet]+'|Electricity|Hydro|dam']
           
                #Energy generation : storage (PHS, battery, H2 storage) (summing positive values for stores and storage units)
                var[sh][sheet_var[sheet]+'|Flexibility|Electricity Storage|Medium duration'] = MW2GW*h*(select_metric(select_time_period(
                    n.storage_units_t.p.filter(like ='PHS').filter(like =country)[n.storage_units_t.p.filter(like ='PHS').filter(like =country)>0].sum(axis=1)))
                    +select_metric(select_time_period(n.stores_t.p.filter(like ='H2').filter(like =country)[n.stores_t.p.filter(like ='H2').filter(like =country)>0].sum(axis=1))))
                var[sh][sheet_var[sheet]+'|Flexibility|Electricity Storage|Short duration'] = MW2GW*h*(select_metric(select_time_period(
                    n.stores_t.p.filter(like ='battery').filter(like =country)[n.stores_t.p.filter(like ='battery').filter(like =country)>0].sum(axis=1))))
                var[sh][sheet_var[sheet]+'|Flexibility|Electricity Storage'] = (var[sh][sheet_var[sheet]+'|Flexibility|Electricity Storage|Medium duration']
                 +var[sh][sheet_var[sheet]+'|Flexibility|Electricity Storage|Short duration'])
            
                #Energy generation : Interconnect (summing positive values for lines and negative ones for links)
                var[sh][sheet_var[sheet]+'|Flexibility|Interconnect Importing Capacity'] = MW2GW*h*(select_metric(select_time_period(
                    ((-1)*n.lines_t.p1[[i for i in n.lines.index if country in n.lines.bus0[i] or country in n.lines.bus1[i]]])
                    [n.lines_t.p1[[i for i in n.lines.index if country in n.lines.bus0[i] or country in n.lines.bus1[i]]]<0].sum(axis=1)))
                 +select_metric(select_time_period(((
                   (-1)*n.links_t.p1[[i for i in n.links.index if 'DC' in n.links.carrier[i] and 
                   ((country in n.links.bus0[i]) is not (country in n.links.bus1[i]))]])
                   [n.links_t.p1[[i for i in n.links.index if 'DC' in n.links.carrier[i] and 
                    ((country in n.links.bus0[i]) is not (country in n.links.bus1[i]))]]<0]).sum(axis=1))))
                      

                ### Heat  (MWh -> GWh)
            
                #Energy generation :Biomass (CCS)
                var[sh][sheet_var[sheet]+'|Heat|Biomass'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p2.filter(like ='solid biomass CHP').filter(like =country)).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Heat|Solid bio and waste'] = var[sh][sheet_var[sheet]+'|Heat|Biomass']  
                var[sh][sheet_var[sheet]+'|Heat|Solid bio and waste|Primary solid biomass'] = var[sh][sheet_var[sheet]+'|Heat|Solid bio and waste']  
            
                var[sh][sheet_var[sheet]+'|Heat|Biomass|CCS'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p2.filter(like ='solid biomass CHP CC').filter(like =country)).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Heat|Solid bio and waste|CCS'] = var[sh][sheet_var[sheet]+'|Heat|Biomass|CCS']  
                var[sh][sheet_var[sheet]+'|Heat|Solid bio and waste|Primary solid biomass|CCS'] = var[sh][sheet_var[sheet]+'|Heat|Solid bio and waste|CCS']  
          
                #Energy generation :Electricity (Resistive heater, heat pump)            
                var[sh][sheet_var[sheet]+'|Heat|Electricity|Direct'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='resistive heater').filter(like =country)).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Heat|Electricity|Heat pump'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='heat pump').filter(like =country)).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Heat|Electricity']=(var[sh][sheet_var[sheet]+'|Heat|Electricity|Direct']
                 +var[sh][sheet_var[sheet]+'|Heat|Electricity|Heat pump'])

                #Energy generation :Natural gas
                var[sh][sheet_var[sheet]+'|Heat|Gases|Fossil|Natural Gas'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p2.filter(like ='gas CHP').filter(like =country)).sum(axis=1)))          
                var[sh][sheet_var[sheet]+'|Heat|Gases|Fossil|Natural Gas'] += MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='gas boiler').filter(like =country)).sum(axis=1)))          
                var[sh][sheet_var[sheet]+'|Heat|Gases|Fossil|Natural Gas|CCS'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p2.filter(like ='gas CHP CC').filter(like =country)).sum(axis=1)))            
            
                #Energy generation : Oil
                var[sh][sheet_var[sheet]+'|Heat|Liquids|Fossil'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='oil boiler').filter(like =country)).sum(axis=1))) 
            
                                                
                # Hydrogen Energy generation  (MWh -> GWh) 
                var[sh][sheet_var[sheet]+'|Hydrogen|Electricity'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='H2 Electrolysis').filter(like =country)).sum(axis=1)))                              
                var[sh][sheet_var[sheet]+'|Hydrogen|Gasses|Fossil|Natural gas'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='SMR').filter(like =country)).sum(axis=1)))            
                var[sh][sheet_var[sheet]+'|Hydrogen|Gasses|Fossil|Natural gas|CCS'] = MW2GW*(-1)*h*select_metric(select_time_period((n.links_t.p1.filter(like ='SMR CC').filter(like =country)).sum(axis=1)))            
            
                var[sh][sheet_var[sheet]+'|Hydrogen'] = (var[sh][sheet_var[sheet]+'|Hydrogen|Electricity']+
                 var[sh][sheet_var[sheet]+'|Hydrogen|Gasses|Fossil|Natural gas']+
                 var[sh][sheet_var[sheet]+'|Hydrogen|Gasses|Fossil|Natural gas|CCS'])
             
                         
             if sheet in ("Percentile_25_consumpt","Percentile_50_consumpt","Summer_peak_consumpt",
                             "Winter_peak_consumpt") :
                sh=sheets[sheet]
            
                #Hourly power consumption : Buildings, Industry, Transportation  (MWh -> GWh)
            
                var[sh][sheet_var[sheet]+'|Electricity'] = MW2GW*h*select_metric(select_time_period((n.loads_t.p[[i for i in n.loads.index if i==country+'0 0']]).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Buildings|Heating'] = MW2GW*h*select_metric(select_time_period(n.loads_t.p.filter(like ='heat').filter(like=country).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Buildings|Residential|Heating'] = MW2GW*h*select_metric(select_time_period(n.loads_t.p.filter(like ='heat').filter(like='residential').filter(like=country).sum(axis=1)))
                var[sh][sheet_var[sheet]+'|Buildings|Services|Heating'] = MW2GW*h*select_metric(select_time_period(n.loads_t.p.filter(like='heat').filter(like ='services').filter(like=country).sum(axis=1)))
            
                var[sh][sheet_var[sheet]+'|Industries|Electricity'] = MW2GW*h*select_metric(select_time_period(n.loads_t.p.filter(like ='industry electricity').filter(like =country).sum(axis=1).squeeze()))
                var[sh][sheet_var[sheet]+'|Transportation'] = (MW2GW*h*(select_metric(select_time_period(n.loads_t.p.filter(like ='transport').filter(like =country).sum(axis=1)))
                   +select_metric(select_time_period(n.loads_t.p.filter(like ='shipping').filter(like =country).sum(axis=1)))))
              
             if sheet in ("installed_capacity", "fuel_consumption_supply","Emissions_supply",
                           "Yearly_generation_supply","Winter_peak_generation","summer_peak_generation",
                            "Percentile50_generation","Percentile25_generation" )  :           
                sh=sheets[sheet]
                var[sh][sheet_var[sheet]+'|Heat']=var[sh][sheet_var[sheet]+'|Electricity']=var[sh][sheet_var[sheet]+'|Hydrogen']=0
                
                for v in var[sh].keys() :                                                               #Summing the variables for the total variable, only needed for some of the sheets
                  if 'Heat|' in v :       
                    masters= [i for i in var[sh].keys() if (i is not v) and (i in v) and ('Heat|' in i) ]  #only master variables are included (e.g: wind|onshore and wind|offshore are already summed in wind)
                    if not masters and 'Biomass' not in v:                                                 #'Biomass' is the same as "solid bio and waste"
                       var[sh][sheet_var[sheet]+'|Heat']+= var[sh][v]
                  elif 'Electricity|' in v or 'Electricty|' in v :                                      #spelling error in table for 'electricity'
                    masters= [i for i in var[sh].keys() if i is not v and i in v and ('Electricity|' in i or 'Electricty|' in i) ]  
                    if not masters and 'Biomass' not in v:
                       var[sh][sheet_var[sheet]+'|Electricity']+= var[sh][v]
                  elif 'Hydrogen|' in v :      
                    masters= [i for i in var[sh].keys() if (i is not v) and (i in v) and ('Hydrogen|' in i) ]  
                    if not masters and 'Biomass' not in v:
                       var[sh][sheet_var[sheet]+'|Hydrogen']+= var[sh][v]
            
            #############################################
            #########  Fill Efficiencies sheet #########
            #############################################   
            ### Efficiency
            sh=sheets["Efficiency_supply"]                 
                        
            #Efficiency : Solar PV (rooftop, ground), wind   
            var[sh]['Efficiency|Electricity|Solar|Rooftop PV'] = n.generators.efficiency.filter(like ='solar rooftop').filter(like =country).mean()
            var[sh]['Efficiency|Heat|Solar'] = n.generators.efficiency.filter(like ='solar thermal').filter(like =country).mean()
            var[sh]['Efficiency|Electricity|Solar|Open field'] = n.generators.efficiency.filter(like ='solar').filter(like =country).mean()
            var[sh]['Efficiency|Electricity|Solar'] =var[sh]['Efficiency|Electricity|Solar|Open field']
               
            var[sh]['Efficiency|Electricity|Wind|Onshore']=n.generators.efficiency.filter(like ='onwind').filter(like =country).mean()
            var[sh]['Efficiency|Electricity|Wind|Offshore']=n.generators.efficiency.filter(like ='offwind').filter(like =country).mean() 
            var[sh]['Efficiency|Electricity|Wind']=(var[sh]['Efficiency|Electricity|Wind|Onshore']+
                                                    var[sh]['Efficiency|Electricity|Wind|Offshore'])/2
            
            #Efficiency : Nuclear, Coal (lignite)         
            var[sh]['Efficiency|Electricity|Nuclear'] =n.links.efficiency.filter(like ='nuclear').mean()      
                                                           
            var[sh]['Efficiency|Electricity|Coal|Brown Coal|Lignite'] = n.links.efficiency.filter(like ='lignite').filter(like =country).mean()
            var[sh]['Efficiency|Electricity|Coal|Brown Coal'] =var[sh]['Efficiency|Electricity|Coal|Brown Coal|Lignite'] 
            var[sh]['Efficiency|Electricity|Coal'] = n.links.efficiency.filter(like ='coal').filter(like =country).mean()                                                  
            
            #Efficiency : hydro (reservoir, ror)
            var[sh]['Efficiency|Electricity|Hydro|river'] = n.generators.efficiency.filter(like ='ror').filter(like =country).mean()
            var[sh]['Efficiency|Electricity|Hydro|dam'] = n.storage_units.efficiency_dispatch.filter(like ='hydro').filter(like =country).mean()
            var[sh]['Efficiency|Electricity|Hydro'] = var[sh]['Efficiency|Electricity|Hydro|dam']   #currently equal in the model since eff(dam)=eff(river)
                
            #Efficiency : storage (battery, H2 storage)               
            var[sh]['Efficiency|Flexibility|Electricity Storage|Medium duration'] = (n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country).mean()
              *n.links.efficiency.filter(like ='H2 Fuel Cell').filter(like =country).filter(like =country).mean())
            var[sh]['Efficiency|Flexibility|Electricity Storage|Short duration'] = (n.links.efficiency.filter(like ='battery charger').filter(like =country).mean()
              *n.links.efficiency.filter(like ='battery discharger').filter(like =country).filter(like =country).mean())
            var[sh]['Efficiency|Flexibility|Electricity Storage'] =  (var[sh]['Efficiency|Flexibility|Electricity Storage|Medium duration']+
            var[sh]['Efficiency|Flexibility|Electricity Storage|Short duration'])/2
            
            #Efficiency : Natural gas(OCGT, CCGT, CHP, CHP CC)                                                                        
            var[sh]['Efficiency|Electricity|Gases|Fossil|Natural gas'] = (n.links.efficiency.filter(like ='gas CHP').filter(like =country).mean()
              + n.links.efficiency.filter(like ='OCGT').filter(like =country).mean() + n.links.efficiency.filter(like ='CCGT').filter(like =country).mean())/3
            var[sh]['Efficiency|Electricity|Gases|Fossil|Natural gas|CCS'] = n.links.efficiency.filter(like ='gas CHP CC').filter(like =country).mean()
                
            #Efficiency :Biomass (CCS)                                                                 
            var[sh]['Efficiency|Electricity|Solid bio and waste|Primary solid biomass'] =n.links.efficiency.filter(like ='solid biomass CHP').filter(like =country).mean()
            var[sh]['Efficiency|Electricity|Solid bio and waste']=var[sh]['Efficiency|Electricity|Solid bio and waste|Primary solid biomass']
            var[sh]['Efficiency|Electricty|Biomass']=var[sh]['Efficiency|Electricity|Solid bio and waste']
            var[sh]['Efficiency|Electricity|Solid bio and waste|Primary solid biomass|CCS'] = n.links.efficiency.filter(like ='solid biomass CHP CC').filter(like =country).mean()            
            var[sh]['Efficiency|Electricity|Solid bio and waste|CCS']=var[sh]['Efficiency|Electricity|Solid bio and waste|Primary solid biomass|CCS']
            var[sh]['Efficiency|Electricty|Biomass|CCS']=var[sh]['Efficiency|Electricity|Solid bio and waste|CCS']
            
            #Efficiency : Hydrogen            
            var[sh]['Efficiency|Electricity|Gases|Hydrogen'] = n.links.efficiency.filter(like ='H2 Fuel Cell').filter(like =country).mean()

            #Efficiency : Oil 
            var[sh]['Efficiency|Electricity|Liquids|Fossil'] =n.links.efficiency.filter(like ='oil-').filter(like =country).mean()

            #Efficiency : Power to gas  
            var[sh]['Efficiency|P2G|Electricity'] = (n.links.efficiency.filter(like ='Sabatier').filter(like =country).mean()+ 
                 n.links.efficiency.filter(like ='helmeth').filter(like =country).mean()+
                 n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country).mean()+
                 n.links.efficiency.filter(like ='Fischer-Tropsch').filter(like =country).mean())/4    
                
            ### Heat
            
            #Efficiency :Biomass (CCS)
            var[sh]['Efficiency|Heat|Biomass'] = n.links.efficiency2.filter(like ='solid biomass CHP').filter(like =country).mean()
            var[sh]['Efficiency|Heat|Solid bio and waste'] = var[sh]['Efficiency|Heat|Biomass']  
            var[sh]['Efficiency|Heat|Solid bio and waste|Primary solid biomass'] = var[sh]['Efficiency|Heat|Solid bio and waste']  
            
            var[sh]['Efficiency|Heat|Biomass|CCS'] = n.links.efficiency2.filter(like ='solid biomass CHP CC').filter(like =country).mean()
            var[sh]['Efficiency|Heat|Solid bio and waste|CCS'] = var[sh]['Efficiency|Heat|Biomass|CCS']  
            var[sh]['Efficiency|Heat|Solid bio and waste|Primary solid biomass|CCS'] = var[sh]['Efficiency|Heat|Solid bio and waste|CCS']  

            #Efficiency :Natural gas
            var[sh]['Efficiency|Heat|Gases|Fossil|Natural Gas'] = (n.links.efficiency2.filter(like ='gas CHP').filter(like =country).mean()
             +n.links.efficiency.filter(like ='gas boiler').filter(like =country).mean())/2
            var[sh]['Efficiency|Heat|Gases|Fossil|Natural Gas|CCS'] = n.links.efficiency2.filter(like ='gas CHP CC').filter(like =country).mean()            
           
            #Efficiency : Oil
            var[sh]['Efficiency|Heat|Liquids|Fossil'] = n.links.efficiency.filter(like ='oil boiler').filter(like =country).mean()

            #Efficiency :Electricity (Resistive heater, heat pump)           
            var[sh]['Efficiency|Heat|Electricity|Direct'] = n.links.efficiency.filter(like ='resistive heater').filter(like =country).mean()
            var[sh]['Efficiency|Heat|Electricity|Heat pump'] = n.links_t.efficiency.filter(like ='heat pump').filter(like =country).mean().mean()

                  
            # Hydrogen Efficiency   
            var[sh]['Efficiency|Hydrogen|Electricity'] = n.links.efficiency.filter(like ='H2 Electrolysis').filter(like =country).mean()                 
            var[sh]['Efficiency|Hydrogen|Gasses|Fossil|Natural gas'] =n.links.efficiency.filter(like ='SMR').filter(like =country).mean()
            var[sh]['Efficiency|Hydrogen|Gasses|Fossil|Natural gas|CCS'] = n.links.efficiency.filter(like ='SMR CC').filter(like =country).mean()  
                  
            ###Energy consumption (MWh -> TJ , TWh -> TJ)
            sh=sheets["Demand_final_energy"]  
            
            #Indusrty
            var[sh]['Final energy consumption|Industries|Gases|Hydrogen'] = MWh2TJ*h*(n.loads_t.p.filter(like ='H2 for industry').filter(like =country).sum().sum())
            var[sh]['Final energy consumption|Industries|Direct heating'] = MWh2TJ*h*(n.loads_t.p.filter(like ='low-temperature heat for industry').filter(like =country).sum().sum())
            var[sh]['Final energy consumption|Industries|Electricity'] = MWh2TJ*h*(n.loads_t.p.filter(like ='industry electricity').filter(like =country).sum().sum())                                                                                                                                                                                                                                                   
            var[sh]['Final energy consumption|Industries|Coal|Coal products'] = TWh2TJ*industry_demand['coke'].filter(like=country).sum()
            var[sh]['Final energy consumption|Industries|Coal'] = (TWh2TJ*industry_demand['coal'].filter(like=country).sum()
                +var[sh]['Final energy consumption|Industries|Coal|Coal products'])
            var[sh]['Final energy consumption|Industries|Solid bio and waste|Primary solid biomass'] = TWh2TJ*industry_demand['solid biomass'].filter(like=country).sum()                                                                                                                                                                                                                               
            var[sh]['Final energy consumption|Industries|Gases|Fossil|Natural Gas'] = TWh2TJ*industry_demand['methane'].filter(like=country).sum()
            var[sh]['Final energy consumption|Industries|Liquids|Fossil'] = TWh2TJ*industry_demand['naphtha'].filter(like=country).sum()
            
            #Transportation 
            var[sh]['Final energy consumption|Transportation|Road|Gases|Hydrogen'] = MWh2TJ*h*(n.loads_t.p.filter(like ='land transport fuel cell').filter(like =country).sum().sum())                                                               
            var[sh]['Final energy consumption|Transportation|Road|Liquids|Fossil'] = MWh2TJ*h*(n.loads_t.p.filter(like ='land transport oil').filter(like =country).sum().sum())
            var[sh]['Final energy consumption|Transportation|Road|Electricity'] = MWh2TJ*h*(n.loads_t.p.filter(like ='land transport EV').filter(like =country).sum().sum())                                                               
            var[sh]['Final energy consumption|Transportation|Navigation|Liquids|Fossil'] = MWh2TJ*h*(n.loads_t.p.filter(like ='shipping oil').filter(like =country).sum().sum())
            var[sh]['Final energy consumption|Transportation|Aviation|Liquids|Fossil'] = 0
            var[sh]['Final energy consumption|Transportation|Gases|Hydrogen'] = (MWh2TJ*h*(n.loads_t.p.filter(like ='H2 for shipping').filter(like =country).sum().sum())
                  +var[sh]['Final energy consumption|Transportation|Road|Gases|Hydrogen']) #There is no variable for navigation|gases , so H2 for shipping is added to the common variable
                                                                                                                                                     
            #Buildings                                                                    
            var[sh]['Final energy consumption|Buildings|Heating|District heating'] = MWh2TJ*h*(n.loads_t.p.filter(like ='urban central heat').filter(like =country).sum().sum())   
            var[sh]['Final energy consumption|District heating']=var[sh]['Final energy consumption|Buildings|Heating|District heating']                                                       
            var[sh]['Final energy consumption|Buildings|Heating'] = MWh2TJ*h*(n.loads_t.p.filter(like ='residential rural heat').filter(like =country).sum().sum()
                  +n.loads_t.p.filter(like ='services rural heat').filter(like =country).sum().sum()+n.loads_t.p.filter(like ='residential urban decentral heat').filter(like =country).sum().sum()                                              
                  +n.loads_t.p.filter(like ='services urban decentral heat').filter(like =country).sum().sum()+ var[sh]['Final energy consumption|District heating'] )                                 
            var[sh]['Final energy consumption|Buildings|Electricity'] = MWh2TJ*h*(n.loads_t.p[[i for i in n.loads.index if i==country+'0 0']].sum().sum())                                                                    
            
            #Total and equal variables                                                                    
            var[sh]['Final energy consumption|Electricity'] = (var[sh]['Final energy consumption|Buildings|Electricity']
                  +var[sh]['Final energy consumption|Transportation|Road|Electricity']+var[sh]['Final energy consumption|Industries|Electricity'])
            var[sh]['Final energy consumption|Gases|Hydrogen'] = (var[sh]['Final energy consumption|Industries|Gases|Hydrogen']
                  +var[sh]['Final energy consumption|Transportation|Gases|Hydrogen'])                                                                      
            var[sh]['Final energy consumption|Buildings'] = (var[sh]['Final energy consumption|Buildings|Heating']
                  +var[sh]['Final energy consumption|Buildings|Electricity'])
            
            var[sh]['Final energy consumption|Industries']=0
            var[sh]['Final energy consumption|Transportation']=0
            for v in var[sh].keys() :   #Summing the variables for transportation and industry
                  if 'Final energy consumption|Industries|' in v and 'Coal products' not in v:             #'Coal products' is already summed in 'Coal'
                    var[sh]['Final energy consumption|Industries']+= var[sh][v]
                  if 'Final energy consumption|Transportation|' in v and 'Road|Gases|Hydrogen' not in v :  #'Road|Gases|Hydrogen' is already summed in '|Transportation|Gases|Hydrogen'
                    var[sh]['Final energy consumption|Transportation']+= var[sh][v]                       
            
            var[sh]['Final energy consumption|Transportation|Navigation'] = var[sh]['Final energy consumption|Transportation|Navigation|Liquids|Fossil']
            var[sh]['Final energy consumption|Transportation|Road'] = (var[sh]['Final energy consumption|Transportation|Road|Gases|Hydrogen']   
                +var[sh]['Final energy consumption|Transportation|Road|Liquids|Fossil']+var[sh]['Final energy consumption|Transportation|Road|Electricity'])                                                                                                                                     
                                                             
            ###Emissions (demand side)  (Mt CO2)
            sh=sheets["Demand_emissions"]              
            #Process Emissions
            var[sh]['Emissions|Kyoto gases|Fossil|CO2|Industries'] = industry_demand['process emission'].filter(like=country).sum()
                                                                                                                                  
                                                              
            col=[]
            for sheet in sheets.keys() :
              sh=sheets[sheet]   
              col=[c for c in ds[sh][1] if c.value=='Y_'+str(year_sub)][0].column
              for v in var[sh].keys():
                 ro=[r for r in ds[sh]['D'] if r.value==v][0].row
                                    
                 if  math.isnan(var[sh][v]):    # sets 'nan' values to zero so they can be summed for regions (EU,..)
                            var[sh][v]=0
                            if 'Efficiency' in v :   #keeps track how. many countries don't have a certain technology
                               for region in regions.keys():
                                   if country in regions[region][0]:  regions[region][2][v] +=1 
                                        
                 ds[sh].cell(row=ro, column=col).value = round(var[sh][v],3) 
                 
                 if country is ['EU','EU_UK','All']:   #sets all initial values for EU,EU_UK,All sheet as zero
                            regions[country][1][sh].cell(row=ro, column=col).value=0
                  
                 for region in regions.keys():          #sums country values for EU, EU_UK, All sheet
                            if country in regions[region][0]: 
                                    regions[region][1][sh].cell(row=ro, column=col).value +=round(var[sh][v],3)  

                 ds[sh].cell(row=ro, column=1).value = model
                 ds[sh].cell(row=ro, column=2).value = scenario
                 ds[sh].cell(row=ro, column=3).value = country 
                
        ### EU and EU_UK adjustments 
        
        #Efficiencies have been summed, they need to be devided by the number of countries
        #that had a value for the efficiency (i.e., only 3 countries have coal in 2040, so its divided by 3)
        sh=sheets['Efficiency_supply'] 
        for region in regions.keys():
         col=[c for c in regions[region][1][sh][1] if c.value=='Y_'+str(year_sub)][0].column
         for v in var[sh].keys():
           ro=[r for r in regions[region][1][sh]['D'] if r.value==v][0].row
           if regions[region][2][v] !=len(regions[region][0]) : regions[region][1][sh].cell(row=ro, column=col).value/= (len(regions[region][0])-regions[region][2][v])

                
        sh=sheets['Demand_final_energy']    #Aviation fuel demand is a single bus for all countries (not just EU27+Uk) but it's added to EU_UK sheet for best accuracy (TODO: enhance accuracy)
        for region in ['EU_UK', 'All']:
          col=[c for c in regions[region][1][sh][1] if c.value=='Y_'+str(year_sub)][0].column
          ro=[r for r in regions[region][1][sh]['D'] if r.value=='Final energy consumption|Transportation|Aviation|Liquids|Fossil'][0].row
          regions[region][1][sh].cell(row=ro, column=col).value=MWh2TJ*h*n.loads_t.p.filter(like ='kerosene for aviation').sum().sum()                                
          ro=[r for r in regions[region][1][sh]['D'] if r.value=='Final energy consumption|Transportation'][0].row
          regions[region][1][sh].cell(row=ro, column=col).value+=MWh2TJ*h*n.loads_t.p.filter(like ='kerosene for aviation').sum().sum()                     
                        
        #Interconnect capacities are added twice (once for each country connected to the transmission line)
        for sheet in ('installed_capacity',"Yearly_generation_supply","Winter_peak_generation",
                          "summer_peak_generation","Percentile50_generation","Percentile25_generation") :
              sh=sheets[sheet] 
              for region in regions.keys():
                col=[c for c in regions[region][1][sh][1] if c.value=='Y_'+str(year_sub)][0].column
                ro=[r for r in regions[region][1][sh]['D'] if 'Flexibility|Interconnect Importing Capacity' in r.value][0].row
                regions[region][1][sheets[sheet]].cell(row=ro, column=col).value/= 2
     
        
     #removes empty sheets from original file   
    for sheet in sheet_var.keys() :      
         del file[sheet]
     # Save file for current scenario
    file.save("results/IAM_{}.xlsx".format(scenario))
